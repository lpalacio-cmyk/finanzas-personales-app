"""
Finanzas WL - App de finanzas personales
Conectada a Supabase con autenticación multi-usuario.

Versión 4.3 — Auth parte 2: Google, sesión persistente y sidebar mobile
- Botón "Continuar con Google" en el login (Supabase OAuth, flujo implícito:
  un JS mueve los tokens del fragmento #... a query params y la app inicia
  la sesión). Requiere el provider Google configurado en Supabase.
- Sesión persistente: el refresh token se guarda en una cookie (30 días).
  Al cerrar el navegador y volver, la app restaura la sesión sola. Nuevo
  requirement: extra-streamlit-components (actualizar requirements.txt).
- Mobile: al elegir una pestaña del menú, la sidebar se cierra sola
  (workaround con JS sobre selectores internos de Streamlit; best effort).

Versión 4.2 — Auth confiable + pulido de Inicio y login
- Recuperación de contraseña DENTRO de la app: "¿Olvidaste tu contraseña?"
  en el login + pantalla para crear la nueva (el mail debe configurarse con
  token_hash, ver instrucciones). Ya no se pierden cuentas.
- Nuevo .streamlit/config.toml: tema claro forzado (los celulares en modo
  oscuro ya no rompen la visual), color primario navy (adiós flash rojo al
  cargar) y toolbar minimal.
- Logo nuevo de alta resolución (logo_completo_transparente.png), centrado.
- Inicio reordenado: KPIs → Disponible real → Ritmo de gasto variable
  (nuevo: gastado a hoy, promedio diario y proyectado a fin de mes) →
  Composición por categoría (más grande) → Evolución mensual. Se eliminó
  el gráfico de evolución del saldo.
- Colores semánticos: ingresos y saldos en navy; naranja solo para gastos;
  verde/rojo reservados para resultado positivo/negativo.
- Movimientos: filtro por categoría; "None" ya no aparece en las tablas.
- Placeholder de email: "Escribe tu mail". Link a wlhnos.vercel.app en el
  pie del login.

Versión 4.1 — Pulido visual y defaults
- Ocultos los anchors ("clips") que Streamlit agrega a los títulos.
- Removidos los emojis de títulos, menú y botones para un look profesional.
- Tarjetas KPI parejas: misma altura mínima, números con dígitos tabulares y
  tamaño que se achica solo si el monto es largo (ya no se cortan).
- Ajustes responsive para mobile (tarjetas y tipografía).
- Todos los selectores de mes (Inicio, detalle del ER, detalle del FF y
  conciliación) arrancan por defecto en el mes en curso, no en el último mes
  con datos (que podía ser futuro por cuotas proyectadas).

Versión 4.0 — Tanda 2C: Reestructura de pantallas
- Menú reducido a 5: Inicio · Cargar movimiento · Movimientos · Reportes ·
  Configuración. Inicio absorbe el Dashboard; Reportes une Estado de
  Resultados y Flujo de Fondos con un selector devengado/caja.
- "Resultado" ya no resta el Ahorro: Resultado = Ingresos − Gastos. El ER
  muestra además "Resultado después de ahorro". El saldo (caja) sigue
  descontando el ahorro porque ese dinero sale de la cuenta.
- "Ajustar saldo inicial" (delta aditivo, críptico) reemplazado por
  "Conciliar saldo": el usuario informa su saldo real a fin de mes y la app
  calcula y guarda el ajuste sola.

Versión 3.9 — Tanda 2B: Tipo desde la categoría + Copiar fijos del mes anterior
- Cargar/Editar: el selector de tipo queda en Ingreso / Gasto / Ahorro. Para
  Gasto, la categoría ya trae su clasificación (se muestra "· fijo" o
  "· variable") y el movimiento se guarda con el tipo correcto. Los reportes
  no cambian. Sin migración: categorias.tipo ya existía.
- Nuevo en Cargar movimiento: expander "Copiar fijos e ingresos del mes
  anterior" — grilla editable (montos y conceptos) con tilde por fila; los
  que parecen ya cargados este mes vienen destildados. Carga en lote.

Versión 3.8 — Tanda 2A: Carga simplificada, filtros y disponible real
- Cargar movimiento: una sola fecha por defecto. Toggle "Pago en cuotas o en
  otra fecha": si está apagado, cuotas=1 e inicio de pago = fecha del
  movimiento (la carga típica baja de 7 campos a 4).
- Ver movimientos: filtros por mes y por tipo; los totales y la tabla
  respetan el filtro.
- Dashboard: nueva sección "Disponible real" — saldo a fin del mes en curso,
  comprometido en cuotas futuras ya cargadas, y el neto disponible.
- Rendimiento: get_movimientos con caché (ttl 30s); el conteo de movimientos
  por categoría en Configuración pasa de ~28 consultas a 1.

Versión 3.7 — Fixes críticos pre-rollout (sesión por usuario)
- FIX CRÍTICO: el cliente Supabase ya no usa @st.cache_resource (era compartido
  entre TODOS los usuarios conectados al mismo tiempo: el login de un usuario
  pisaba la sesión de los demás). Ahora los tokens viven en st.session_state
  (individual por navegador) y la sesión se rehidrata en cada ejecución.
- Esto también arregla el vencimiento del token (~1 hora): set_session
  refresca automáticamente y la app ya no "se vacía" en silencio.
- do_signout ya no limpia cachés globales (no le rompe la conexión al resto).
- FIX: en Ver movimientos, el selector "Tipo" del form de edición salió del
  st.form (los widgets dentro de un form no refrescan): al cambiar el tipo
  ahora se actualizan las categorías y ya no se puede guardar un movimiento
  con categoría de otro tipo.

Versión 3.6 — Ajustes finos sobre Tanda 3
- "plata" → "dinero" en toda la app
- Estado de Resultados: removidos los gráficos de evolución mensual y torta
  por categoría (ahora viven en Dashboard, una sola fuente de verdad)
- Inicio: "Saldo actual" usa el mes en curso (no el último con datos), para
  no mostrar saldos proyectados de meses futuros como si fueran "actuales"
- Dashboard: gráfico de torta rediseñado con filtros propios (selector de mes
  con opción "Todos" y selector de tipo: Todos / Gasto Fijo / Gasto Variable / Ahorro)

Versión 3.5 — Tanda 3: Pestañas Inicio y Dashboard
- Pestaña "🏠 Inicio" como default al loguearse: bienvenida, resumen rápido,
  guía de pestañas y aclaración corta de devengado vs caja.
- Pestaña "📊 Dashboard" con KPIs del mes seleccionado, evolución del saldo,
  evolución mensual con selector de serie (resuelve el problema de escala
  cuando se mostraban Ingresos vs Gastos en un mismo gráfico) y gastos por
  categoría del mes.
- Refactor: cálculo de saldos encadenados extraído a calcular_estado_flujo()
  para que page_flujo y page_dashboard compartan la misma fuente de verdad.

Versión 3.4 — Tanda 2.1: Saldo inicial como AJUSTE ADITIVO (no override)
- El monto en saldos_iniciales se interpreta como un DELTA sobre el cálculo automático.
- Fórmula: Saldo inicial M = Saldo final M-1 + Ajuste manual M
- Para el primer mes con datos: SF M-1 = 0, entonces el ajuste = saldo inicial de partida.
- Para meses intermedios: el ajuste reconcilia discrepancias y se preserva en meses posteriores.
- UI: el form muestra "automático + tu ajuste = saldo efectivo" en vivo.
- Cargar ajuste = 0 borra el registro (vuelve al cálculo puro).
- Fila opcional "Ajuste" en la pivot, solo si hay ajustes en el rango.

Versión 3.3 — Tanda 2: Saldos iniciales por mes (persistentes y editables)
- Tabla nueva en Supabase: saldos_iniciales (user_id, mes, monto)
- Saldo inicial de cada mes se calcula encadenadamente desde el primero
- Form para ajustar el saldo inicial de cualquier mes y persistirlo
- Bug resuelto: el saldo inicial ya no se borra al cambiar de pestaña

Versión 3.2 — Tanda 1 de mejoras UX:
- Formato monetario es-AR ($77.569,90 con punto miles y coma decimal)
- Separadores es-AR aplicados también a tooltips y axis de Plotly
- "Fecha devengo" → "Fecha del movimiento" en formularios
- Inputs de monto vacíos por default (placeholder "0,00") en lugar de "0.00" preescrito
- En Flujo de Fondos: "Flujo neto del mes" → "Flujo neto", "Saldo acumulado" → "Saldo final"
- Eliminado el gráfico de "Flujo neto y saldo acumulado"

Versión 3.1 — Rebranding:
- Identidad visual WL HNOS & ASOC (logo, tipografía Open Sans + Poppins, paleta del manual de marca)
- Login rediseñado con card central
- Métricas como cards custom
- Charts con paleta corporativa
- PWA con logo real (icon-192.png + icon-512.png)
"""

import streamlit as st
import streamlit.components.v1 as components
from supabase import create_client, Client
from datetime import date, datetime, timedelta
import hashlib
import secrets as _secrets
import httpx
from urllib.parse import quote
import extra_streamlit_components as stx
from dateutil.relativedelta import relativedelta
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import os
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as _XLImage
from openpyxl.utils import get_column_letter as _col
from openpyxl.worksheet.properties import PageSetupProperties

# ============================================================================
# CONSTANTES DE MARCA (Manual de Identidad WL HNOS & ASOC)
# ============================================================================
APP_URL = "https://finanzas-personales-app.streamlit.app"

NAVY = "#102250"        # Principal corporativo
CYAN = "#1595BC"        # Secundario corporativo
GREEN = "#1C913D"       # Apoyo (positivo)
GREY = "#6C6D6D"        # Apoyo (neutro)
ORANGE = "#EA5E2D"      # Apoyo (alerta)
RED = "#C0392B"         # Solo para resultados negativos
WHITE = "#FFFFFF"

# Colores derivados (para fondos suaves, hovers, etc.)
NAVY_HOVER = "#1a2e6a"
NAVY_LIGHT = "#e8ecf5"
BG_SOFT = "#f7f8fb"
BORDER = "#e5e7eb"
TEXT_MUTED = "#6b7280"

# Mapeo color por tipo de movimiento (para charts y métricas)
COLOR_TIPO = {
    "Ingreso": GREEN,
    "Gasto Fijo": NAVY,
    "Gasto Variable": ORANGE,
    "Ahorro": CYAN,
}

# ============================================================================
# PLANES (FREEMIUM)
# ============================================================================
# El plan de cada usuario se guarda en la tabla `suscripciones` de Supabase
# (RLS: el usuario LEE su fila pero NO la escribe; vos la seteás desde el panel).
# Sin fila => gratis. Helpers es_premium() / _plan_usuario() más abajo.
LIMITE_CATEGORIAS_PROPIAS_GRATIS = 5   # categorías propias (total) que permite el plan gratis
MSG_PREMIUM = (
    "🔒 Esta función es parte de **Premium**. "
    "Escribinos a WL HNOS & ASOC para activarla."
)

# ============================================================================
# CARGA DE LOGO (cached)
# ============================================================================
@st.cache_data(show_spinner=False)
def load_logo_bytes(path: str):
    if os.path.exists(path):
        with open(path, "rb") as f:
            return f.read()
    return None

@st.cache_data(show_spinner=False)
def load_logo_b64(path: str):
    b = load_logo_bytes(path)
    return base64.b64encode(b).decode() if b else None

LOGO_PATH = "logo_completo_transparente.png"
ICON_192_PATH = "icon-192.png"
ICON_512_PATH = "icon-512.png"

# ============================================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================================
_logo_bytes = load_logo_bytes(LOGO_PATH)
# Logo (símbolo WL) embebido en base64 para no depender del archivo del repo
_LOGO_PNG_B64 = "iVBORw0KGgoAAAANSUhEUgAAALQAAAC0CAYAAAA9zQYyAAA6z0lEQVR42u19eZxdVZXut9Y+59xbc1VGAqGqEgYx2Np0EFGRQsYMJChQhcioIHbb+rpt22fr83VRdre2ra3PiSGAiIJiShAxE4OQUkQUsEUkikKSCoFAQmoe7r3n7L3eH3vfqkpSSZ1bqVt1K7nr97u/H0VqOGfvb6+91trf/hZQtKIVrWhFK1rRila0ohWtaEUrWtGKVrSiFa1oRSta0YpWtKIVrWhFK1rRila0QjEqDkHexpWAZgCbRhnjnYQG959tc2T0X7FIgJbsv0lxSIuAnoSxayYLWAfQtjMM0GLy9rcaNvLwAmg1RaAXAX2QY9XIaNhJYwF3xowllbpCZhB5VSJRDXFQLkbPYpVIAIARXSoSVQLQDL+blEpDItY66lag7kjCLiOJ7oT0d3VgoAPbnxjc/2M1qhHPJIc7yIuAjgXgjRqgvYCyKKg8uvZoKHUsK+940XI8mOoJMl+E5gBcTiTlRMwgjjnc4uCoIUZCAL0C0wvQKwTaJqK3gOgvrMPnUxJuGXzp4VeKAC8CemwQ2y1dj/yH0rplR3jEbwb4FCZeLCSLADqaSZUMA9biR0QAGcKScV/shV23QEhGmQNi+ywEEIHAANn/hghEIkCkAyLtAD1jED2FEL/xBqLnOzo29OwD8OF4XIqAPkxBPHPmOyqi8prFYHUmQKcT8ZuIeCZYWXyKgYgBYAwEApA4cJIbVjqAC97fVIzyMzKcGEr2bxADRMQMwD66mAgAtgLyNERvDMOorX/7+ufsgspagwfkLc4vAnpqrZltktUWZf9PxZFnzeRkaQOBlwN0BhEvBHkADEQ0IEYgpEcAd8T4idnL87ID3h5YpVGGXUaCfNizu1VDYr991L8pQz9OpEAMIuWeJqMB+j1BHgD02s6aHb/G00+Hey7iQy+xPAwBPbQFWwDWNSQrVPmZLN4lRHQOsTcPoBEAhh72iiPAO+wtHZB4uFqXxaIIIDoUIENE3QAy9ke5h4hCiJAIkiBTRlACkXIhU0aggEgp+ydHhDNDu4LbIkACCA+DPAtwEhA8IgWQgpgIBHnWAPcjCu/t3rbmt/sdjyKgp1NYMSKkqF/5hojo/UTURKxOANiBWNsQgkZ6wyFPqUBMNAQ0AzGRCGgHiXlJgC1M2KJN2M5Qr2qjdyHSuyMO+wKl+4MBRLu4U/Da7zMA7LMce6yP7qP8uRxIP2dKfEqUIKmqjeaZQjyHROaDVS1EFgJUR0y1IKom8tybGIjJhj5k3M4wDHAhsye4M0aAXwrkLjOw+96+Vx/bdSgBmw4zIHNN3YpzRakPA1hKyk/AGIhEDsRDblYgYqzDZWXBwDZOFb0DQpsA87Qh+Z2W8DlPq5e6t63tnIT38Uprz5vN4IWKvROJ1N8AslhAb2BWFSB2XlwDItqFOu6dxAyBmz0ABDGZHQDupjB1W+dLDzw3FIoBmK7APoQB3aiGgFzXkKxR1RcK1EeJ6O0gcuAU7TwxD+3pAIEU21hUICbqAORpgjwixvxKZwb/0PvKz3bvd/E0Ati5k+wByCIXn7ZIbnOSPWHMHtgMHabo0X6gZuH5tUbjb4S5gcHvAvBmUr5vn18DYkaCezjeJ6WIPIgJUxD9EyL5Rufmn/5yOgObDlEgu2RnUVC1YOGVBO8fSPlvshMcis38R3guQABWxJ7dpU30ioh+lJnWpNKZxwZf3rB9n4SycROhFZjkkpibr2YCNjIwR0YDeXntkkWenzgLQssFeCezX+4W576LWMSASBH7EIkEkPtIoq90bl7z2L7jWQT0ZBpbz2Y9Ss3Cle8z4H9h9t9it+FIuyrBSG/MRB6BGMZkuiHyAGBWS/j6oz3bn+jYtyIyRwp0chloJDQCaN0T4JVHLzmGVWKZEF1MhHcSB0okAow+ELABwT1C4ee7X8wmkA0e0KYLHdiHCKCHw4uquuVnkAr+lVi9GxCIjvaaOGgbR/o2rDT6KSF9F2Pw3s7ND2/b83cC07O01czWg+9Zc66qW3kSMb9PmC5h8urcQnfv5+p9EA2AiQMSiUIYuZ0o+o/OzWu27RPKFQGdj4mz233NwvNrRbwWMF9NxBATuokc6YFsWCEmzADyE4G+uXvzmp/t+fs20aFVn82GR8PvVFV7Wg3UzAuJ8GEi7602pxgKxVwhW7T12AHEhLtEoi92b3nl68DTYSGHIdMY0A0e0BYBoOoFKz8CUs3E/mwx6VEmZgjIfRC5ExLd1LV17TOjx92Hsu1zmETVC1YsB/HHiNS5tnS5tyOABpNH5EEkfEqi8JPd7Ws3Fqq3no6AZnuAQFJdv/wtUP5XiPwzbcJj9NBRGUQDLiY0UT9BvhNJ+uu9W9b/eS9vrHH42T51+aoFK88mqE8S87n2YCncJ+cg9pQYEYj5WlfPzn/F7sd7C80ZTDNAj4iVF6z8OLH6dyKvVHRGD9eQRQAYYl+JMSIw31M6858d7ev+OPw7Dp2TsYkZ09UmyyasXnDBChB9htg/VUQDxkQgeMPlPiLigIwJn4POfNR5a8cNn/oxpekG5ppjzj1aUHIzUbDUEnKMHg4vEBGz5457N8Lo5q6t9/+8COScga2qFlxwNYj/ldmvFZPJgtkxoSQi9jxAtDH6P7q3/KTFhnlTH4JMB0APrf7q+pUrwepGYu9I0WEEgsJwLRnEAYuJXhaRf+3ect+3i0A+uF2w/Ngls5UkPkvgjxIUiwlH7oTWW6uARGcehslc27V1XftUg7rAAd3MWSBWLVhxPbHfDGDPbVBEgz1FAMREt6ej/s864vs+cWLRxhdjly88/zQP/n8T+6fs5a0FEE3seyL6VR2FH+xtX7N+KuNqKnRPUVXXUE1qxq2k/ItEZ4wdo+FSHKlAwUTtRuuPd7ff/+O9KiBFmzBgLwqqFhz3GSL1GSLyxUTDCbiIBisFQESiT3Vvvv9LGKYeTurOqAqzItfgoX2drqg/7w3kVaxn9s8QE0a2wuEYOGAm5bPo8O4QPRf3bt3wlF0EmwC0F8OLCbNNYse1LUp3Pb/Rq6h/hBWfQhwcAdHaXqkhBowBQMyJc5PVx81PLSxbjx07tJsTmcQVWHBo9oC2qLp+2elQJXcT0TwbL9NQiEHsKREZIJh/7tx83417x35FyxdWGhTQFs049pRKY478CrF3jRgNC+bhujV5gSc687DJDFzSs/3BjsmcGypEMFctXHEhkXcXgZJivUB2J4mIfc+Y6I8ig1f1bNnw5MjTwiLmJjdprKpfcR0p72sETu4VgkSkEp6Y8H8k6ruoe9tDWyYrDKSCA/OCFdcw+6sEwpa5PqKwrxLK6PRPoLuu7m5v6yrGylMfW9fULn2neMnvEnsLxWQiYOjmQUTseWLMi2Gm9z392x/6w2R4aiokMFcuvODDir2bxGh3mZSGy0Psk+jwi11b7vu0BXgxxCiUeSutPW9ewiv9AdhvEJ0ZcRCDiEh5InqXkYHze7Y8+Jt8zxsXRAKItqhy4Xv+TpF/kxhtRoKZiJmIjejBv+/act+/2Jo0qAjmQrC2CGhUA9se2NGZ2bHUmMxdpALPMfYEgCcm0kRqNlPp+soFS95q560xb8UINeVgbmuLao55z3VM3k2OykhDnpkUQ2hAJHNZ95Y137UDcUMxXi64Kkgzo+euMN35p3uT1cdWEgfvhGgzVAERo4m9MoJ3YVA+/5F0z5qX81X9mEJAN3hob4uq6lZ+kNi7RUQbe1U/C2aPCeiESa3o3rp2g93e1hW9cmF6arGOqJlTnTc/kKg6hpmDd9vqh2TLeppIlYH9lYmy2p+ke366Ox+gnhpANzYqbFqnq+qWv5e94E4RIyPBTOQxxHToqH9FT/uGx4rJ37QBNoBGle768SOJmuO6mf2lTneB3C1jzcqvJPKWc+Xs1rB7fY+tUrVNGKinICl0JKPaZe8QP/EzCBKAkeyBCZFiEek0mf4lPdsf/E02LCmCZRpVQBobGa2tumrhBf/I7H9VTKRdvkYQ0aQCJSb8tfJ3nrP7+cf7XM40IaCeZA/dzMANpvLoJcdQULIBoJq9wQxQl9b9y3tfevDXRTBP17B6k6ChwUv/7oHHE1XH9LEKluwVU0fEQa3WiRPTnc/fDWxie8I7vTw0AYIZx76tQpt5v2Dy32xJ5KTcUTaBKE2SOadz85rHphTMIoTrrx/f2LRcLxPlbaa9p3Yni1ULz/88c8mnHUMye+IbkZfwRA9+pWvzTz8xUWHl5AG6sVGhtVVXL7xgNamgUaKhl7NXpoiVhIMXd29bf0/RMx+KoF55A3Py7+zhi5t3gSble0b3X9u9Zd1tEwHqSQK0jZsr68//rPJK/010JgTBd/8YEfue6MGPdm1Z861CSADnfurG+oyoMkK4H5JTAAR7/a8MIIoU9b20reMbLb1ZX1/E9PCpYtWCFT9hL7lSokyWm2NLsMRpYzINPVvWPDmSMjwe8yYLzDPqzz9XVNAiJqP3OEniwBOd+rIF82IfaAunbOhXi0ITaVMz6xp/xrzPymBfZviO4li+QDSVlAbSPetmAH+L1asVmpqKZUbAaZkIqe63XWGqj9xIyj/J3sontvEdlTCpO6vqGt7W3Y4eDCle5m55PilsZqDVlNaeN8+wf7v9e1lZWEsMNya9tmvL/Z+0wH96asOM564XAMgM9twuYXoAiUQAz1Ojf9SeH6UCGAESJY3ln/7abDQ1aYgU9beHQN3EHR2/6ZEovFiM3kWksrQGFhNq4sTx4KobrHduHDcu8wzojQxAfC+xitg70jHnbEWDfSU6fEEyg1dagLdO/QlgS4uBCHe3XLMZqYE1FJQItNYQESutO/Ije34AkUxaU1nljNKKIy8CAFy/URWxnDV75N29be1m0uFVjv9PLuRQYjIRq8SlVQuWf/BgjsfzCGhLCq9ccP5HmBPniw4jRwMVgEnEdJowvMzxZa02bUGMeysBgO7puQVhhsA8Qmh85Af7fogAIzAlpVcAYFx/RjHk2BvUDQ1eZ/ua9dpkPkukZMS8KzGRAQVfqTx6yTEW1M1cIIBuZqBVVxx51vFMwX/aBx2qeWtij8SED/ZsX/ubgmPNNTUZiFDHpvseNYN9v6Mgye60K076oyQ9IJwsOXXGZ245GUSCZuEikEdYW5sFarjrJhHdg6zMK0CAEWZVRX7iJvv1JioQQNsHUcmybxJzhT08GcqilIgGiBvKjzhttgNzIcWagus3KrS2akoN3EbKy7FaIZqSZcwz5l4BADix2Mdmr52bgRZD/pz3EPvVMGbE/JMSE2pWwdlV9Ss+NJ7QIw+Adpdba5ddQ5w4R3QUjbhxYv2YaM0cHKFKqlYMv2QBWYsNFQZffv6HZqBnFzxfuTg5znJQJpMCBcFFVVc1V6OJtEuEi2bDDsftwDVWI3LffU5MZIjVF0qOWjLffn/80GOigcTAalO24My5pPzPi0RiWySMPvMEZb3YkDB4oRgJVovq+8b/2SXpgR9RogQgxAuLiAhh2lBZ1Txv0Yl2wa5GMewYcnbAjPplpxCpt4uJZC9nB0diEmJ/RhAE/2V3x/ihxwQPdCMBJB7KP0cqmANjzJ7NdoYeWomJBKTeWbZgyZttqaa5sCa9tRUAEO7uuFVS/dG+A3/AYRUAwkHp1bYcWDxg2TMm8z5I7NH+cxNSYjKa2bu0qnbpmbmEHhMIIpsIVi5cejKRukZMaPZ/KJFNDn3fQ/D+kXF34QC6SaO5mbs//8HfIjXwc0qWESSulwZLegDwk++qab71TWghg+bmw91LE9Cqy49dMptAF4lEcCpM+/9+YpDy/wuLF/txd/EJH2Q2weeJlRq1e+qeC5VFIhCoCXPPKSvA5BA40RKUJN1/s+uiRrEnzxjDZRW+Kp91xcjfdfhagwIA1l4jecFM2/zlQPPtDlxUYnHV60dd4Q5c1CQBulEBLaZi4bLlpPxzxIR67C2aGKINsb+gKpk4pyCTwyayR7Z/+M1a09ezhfxE/Ho5EUmYBgXJpnnXNZfa5PBwrnjYch1DfQASNwIjEjFCiv9v+fHnzxqRUOYV0GS3gwZPGf/6nEtkRCAu1OQQgtXg1+78734JU98lP4GhDlJxAulM2lBZRX204I1LbXK4+jANOxoVAKmuf/I0sFpsuwXEykkYEgmxX6/SOM0miAd2ehMwwLauWLWg/D2k/JOHOc6x1gKLiQCi86rqzqt3LKvCmnTH74h27/iu7usZACsvh7K0gBkIyj5gf1fjYZ0cCqlrLIcj9qmwASkyJrMN0tXmkhuTZ0CvNli82Ceofxm1kfVY3l0kIvbLQEGTXR+NhbUtZ/kd//mRzZRJraFkKQBEsRdsalAkkTyr4pM3HmeTw8Pu5JCAVj3z6LOPBKkLRCKMXvkabQWIEHkEMd/p3vZYJxobs6eK+QJ0owJIqnfPXUYcLB6mBObyusK2wM7vBxb7rrlNYZnjd5i+rlXIpONPCIFgtFFllcnEnHmX2uTwMIujG22IEHkllzL7VXueDI4VjjIbHfbCZG4HQGgdOyQ9SEAvEgAs5H3cPeI4tlRiMZEQq7dU1897B4ZUkQopObRU0N1/uH+jDPT9DyVKcuF3kIQZkOdfjoarkmiEweHUY7211WDxYh9EV9lKUex318QeAfre7vYHtmZD2zwC2lY2ymuXvp3YO330U5/Y0ZUh8gDmKwp2ahy/w6QGbiOlcli8xBKmDJVVHjfrrHPPBJEcPslhIwOQyo45ZzD5f2WFhGLv4EpMZASpG3IpGBz0wCrP+wiR2vvUJ0dPTWy7l/IF5UcsmV2QNWlHBQ12tf9Q9+XM7xD4Pqi86mq3DR9WySHDv9bhOOZ7iyb2yIhps3p48a9l8bif0TXwIagV+576UK5gJBitSfmzVIl3gf1fDYUVdpDld+z4yj+/jvRAKyVKEDtbJ7CkBiB+sLT6E1+qAx0WJ4f2HmHdefUgXi4mRA7emQCATPhN++XGfJOTGhgAjCTeTyqo2PPUR7TR4dcgSOeaC9sFnK1Jn1GAyaEtxFBv5y0y2BeNcXS7x2qA0ZEqryr35x7b6N7vEAe0rVYJ+5cRB2UQRPF2XTEgj4wO/9jtm7X2Z9p0ngHdpoHFPoQut0KTcBJeChB6vqf92X8Bye+JPYl/EGEJS0T89ur65W8pUMKSRrPwrpYP/E5Sg2058TsAliiE+MEVuO46/xC/zWLbS9c1JIn4StsDNXa5UmytWt+KFzak3U4teQS0rQWW189+B7F3opW/daLkVjPmB0B7yoj5LkAEyYkcb0h5Psi7zH69qfCqAdmy22D/KrtW4/I7iCWTMlxa/uaZ8059F4iAxtWH6J1DmwxWc8U5zMHxYmIng/Z+oc68Hg2k7xx2njnFObk+q0sG4Tfaojdc+y5WxqQHje79AQCYwfCHojM7s92R4k46jIYQCpew5PgdfvvP1pi+ns0UJDj2LiQiCBKgyllXApDsWB56ZisSQupaex0zdjJoiDwI9Pf7X3toZ9Z55hPQhNZWPXt2QznALhmEcn2gQWIe7Xnp0RexeLHf9+qGXSKymshD7JotQCLaMPt1VWXJ80as9kIyQfNGtWPVqgGE6TvgJ5DDLsSSTgF+YmXpdf8+D02WonpogdlWJCoWLj+OQOfabr+xyrkCMIvRaa3DmzHEEco5E815K0G6rOId5Hm1tgfKsAfVIncDIAwudHfU9a1iwmzyJLEBAwKErxy52gvLNhoACF/f8V3T39MPFYcuC3ubJQo1l1fVlCx4w4WHZHLYYCsSLHwlsZ90BYM4u5ch9kgk2tC3bcOm8fYOz20wG4eqUCsIjJHhhphMh+H0AwAEm1ojANS1de0zYvSjxB5ySA5ZTAQinFtVu3xhQRKWWloMmoW7v/h3WymTWsPJsuxYxMuXjAaXlB2KUgeEtrZo7txzygh0ub0MHbOESyB7kmi+cTD5E+f0sK22oyiIznaEd86GGxD5ef/mh3a6yoTJlvZA5mbIkHZF7BknTpQIUdPIMmFhJYetBIDMYPfNkk7lMnFK0oNCyZK3zmi+/ZRDS+rA7uCp0mAZqaDeSujGwZgYIp/FRE91by7ZiIPooZPDQDYTAFTVLvgrIjpeRMvIzFUga+yDZIvgbRoAdad2rBeTeRGkOAcuMUE0iPkyp3dXgMlhk4YIdv/uJz+XVN9vKVFC8XMF0ZQoZa6ceeUelZNDJBkE1DXD4WOsIFNABCF9gxOjGXf1JwdAO6AqPoPYz4IzG24MQg88ar9uM8Mv08jY8fQABN+xyWEO/AfRwuy9qapubixi95TYEL9j8FZiRTlMoDLpQcAPLqy8pnkGmugQ0MGzyWDN/HPeRKTOdNyeWN4ZrJQxmW2Ko3tc2KInAdDu5E7ojD23CgUBnune9rMt2KdZuaWCphB9z5hMP5hzKcNokAKRd2XBzqHT78i8tGO17uvZSV4Qj98xnBzOTRx34gVDi2Nam415xU9eRcr3AcSmiTpn952OFzb0ZGvY+QY0AS1mxrGnVIL4JBENeyeQ7I0MkTYAMspWIUCjSm1d1w5j7nftuuOuPiUSAUwryuaeM6cga9JOv6P3Wx/bjXTqh5RIxtfvANnCSLLsyj0cxnRNBtGqZ8w4pRLg99lSHeIdpIDZmEwfp/pvw7BoJ/IMaLvdh5nqNxHRURAjbj4ZomEo+oUNm+fIAcLGVe6YPP4iMkYTBzO9kuR7Rj5HQZnV76Coe9dtZrBvhCBljOQwNSAIEqfN/uxtb0ELmel7cmjnxVTNXUkqmA/ROnYyyD6RmB917nh420SIdsYDSMNOq1XHicW2opH1QsxGRz1+JvU7N7ujTKT1rN3ta38uRj9F7LHrNBpr3btFfIX9anUhEpY0moW6/v2aZyQ1+CiVlFH8Ep5oLqnwUD3bErIap+vR4SLXxoquzS1asFIFBvrGiXqSeIB2nlcIJzvPTIAIMYNAf+zY/sgr7jv3M5ENCoAxom/N8cqhJSyB3l5Vt+SvbTOe5kIt4YEGB1bB5MbvMJk04PtNMy77WOX0lDqwyWBV7fknEfFpOSSD9kaK6LaeLWt/A0c3nQxAW+YUGhWB3+QOxAhCBlAQyP9gzGtTLmuNMq1Gp3eCcuF3iCEVKCL/8pHJR4GV8AwgtKv94XWmr+dFCpJxS5SMMG2ovPJoPuEd01TqwM0H8weIfdfRLNaiJAggJvqWC1smZF5jU/pK63bOBtFCG20QOS8NEe3CjZ104OC/UVlxc7o7t5NDkJgIwnzx7EUN5YWZHFp+B1atGkAmdQcFQS78DgGxoKzyagDTTeqAnA74TCJqcrT4mAcpHhsTPt/t/3ntsNOcFEDbA5UkSupBVD1cliIWoyHKPGu/nhNnIogk+rboMIx/e4HYyu/6tZnB8iUY6qpUaOb4Hbu3fk/39vRBqbglPCWpAaJE4t0zP3nrG6aXDp6dB06UvJdUMMclg3FvdANG34wXXkgfbKkuR0DbAxVD3nGuE5fJzgRguqJMZqvLjsZYYa0aaLb8DtGP2Bu9ErfE5QhLytIuD7K0kxdraTFoXK26v/jxrRQOZvU7YtJKTcRlFQmeM8vywKeNDl6rO9rmDwIiMfdNV6oLX9d+dOdEeud4gG5wT0F0rAUVCSBCxBAxOwZf6tyVgxezf0+bW0bWMWKYEhMBzGfb/hsoPMISMETeMr0dVr+DYj8jS5iB+In3o/HjJbiEpwFhyeZMNXXnvY1YnRqbJirIJoPf73thw66J9M7xAJ2tLRMfY5md4joXEQi0FXg6dJWHGA9lORml4Y71RocvgLy4yRMBRjP7JfD8SyYyiZjg5NDpd6z9uQz2P0WJUor3fu42S1n5MTPfcvLZEAFWF3hN2i5eEZW4xl30iFuKVWKiLOcZE00PjuFBsrVfqR3CrNgQgEi25lh5EKBB7djx9ACJ3EGci76FpRcSyCosxVCinBLL8jsG+m8Dq/jJoUCgfFCpSw4LW+qA0NqqyxaeMweg9w5f9IhTqvMJotc7zvNBdY0dD6AJIMGiRQGBZg+V7FwIrbXZZsOSnTkAK0te6v2e6HQfEJffQSwSGWbvxKraeaejYAlLjt/x8l9aTV/XTvhBXP0OlvQAKEieN/cTX19Q2FIHluLgm+QlxMGMsbWeMRS3QjRE62/l6AgntmxXnqmtAqjGHoC5kp0YMPEre4Ql8cwAjapr66PtYuR+W8KLzX8wIAWoAlZYoiy/4zOW3xEk4yWHRAStNZdVlpm5811YVai3Wdpsiwjiq92rxZEn0KQ8EqOf6t7WtzHb8WGSAW1LdiqjKgCpHOFHSURgRHYe3L5lVklu/A7lVP9XzFmwYm6B1qSH+rNwx65bZaA3/hU0IkgUQoLE5Vg8JHVQYO9nL65W1Q28i1idNOLWf6zVriE3Am1Rvg7IYj0IaV0OUMnwnBABBtDpjvEF9haIXVvXPCYmzIXfQRCjSQUzQqL3jtz+ps5GCQssv4N3fuFDv5f04KOULIt7UVhJelC4tPzEGStPO6OgdfCUuga2X3ecuReQ5Tz73a/+aKJLdTkA2nFcFc1yi9AMO1dJA+gY/0JvZAAagltyd0ICEbrcPv/GKS5x7S+psSVKGui/GWIotuomQZOfgKqoKUQeOAGtuuSoJfMhvFIkLm8D2nKezXc6On7TM9Glupw9NPuBTzTyQgYBJNoTTrlJHc+2bABAovSPxISvxed3ZFvC8alVxyw/aYoJS1x59HlNQIO3L87P0BChXU88st70db9AfjIeNVKgTGoQ8IMVMz7+30dlS4GFEW3YJDzwg0tZ+ZW2bd+YC1UAUsZk+kj6vu0mP28VnFhA0FFUsweBzG40GpQ5mAcb4neI6B/mRP4XaFKeIqMsYalx0glLDABVtWfVs196Q3ldxbF2VEYuLLItltesGkA6dQf5AWLfZtGRpvKqKp5bf9FQKbAwSnUGixYFILo6ttbz0EGKuadr66PtLgY3UwpoIlQ4QVFxKw5EqltzycAwOMdjNvaOMuFtYsIMABVzaJUYDQFfNHPmygp7G30ykyd3C51LG5VfOtNT3IRRO55afofu2PVd3dfdH5vfARC0BiWSV6KxURWG1IHTeu6vP5PZWxRb69kepGgJM98aX76VB0CDPDNKFGvYSx/kSrOCjP0vb/i9iN5I7FP85FBrVt7RYYUsHTHgkxRHtum6uoYkiK82OiMAXTV3NOmylhaD1atV5xc/vI2izH2cLIt3RYuIJTUglCz5m9lvuuBUlxxOsZd2QCT/mvhaz1mdZ93Ws33Dk/k4SBkXoNXekcLBOOX9JE+AWeV+Z2yKi93zsgpLk0VYsp6q2ys7j5V3guhMROwvHEwGKw60sKRn9y2SHpQc+rNoJMvIlFVeVQDhhiPxn7WQmZbF1np2dCWiKG8HKePz0Minc7D8jorw9XWiwxesfkesgwjXm4XOrDjq7OMwaYQlJ0Qo/ofcnxMAYFbXjrqluqTu9earfmEG+5+iRAnHyxWIJT0ISiTeM++fvjwLTU06/k2YCV/EFphccjlxUBpP61k02CNjwj/NZjOhnOeDTwqR1xBOgAa1ffsTg4B8x5H/48WZMIY4KCG//JI9Ytu8md0yy+uWvZHAZ4sJBUS+rbp4DVV1S04aVdfaJnUGmcFbyd6Sjzc3Uai5rHJ2OPNoW3Nf3ToV1RwLxPmnloDU5ZAontazAERMEL3qhRc2pPNZqhuHhzZ7vh/iRwYxvbQ9PyXzPaPDXoDjvrxVWCLzfixaFORfYclumcze1cR+wh2WkM3klQcObIPNhr1aKGT5Ha+2/0j39bwan9/hzi1Ky64EQGhsnIJLwhaIVWrWecT+cTFPBgXMSnT4uk7hey4knJRnjwdoE5l94UVsosREeQwDNKrOzWu2wUT3W/J/zG1ZIsPsv7Fy8JgG5JewZBtIzlxZwYT379FXhpyGCOiSiiPPmom2tj23ZMfv6PnqJzook747Pr/DSh1QovTts5q/d9LU6OBlk0G+BkQSS+tZxBD5EOi7+naseX08Os/5Lduxtws0REIhQADRVcoMlk60u9bQWX5H3MBdQAwSyvPJmj1i1xX6vcSJ+XtdNyKI0ayCORSUXzxqcvgcBADR7tdul4G+ML7EsGguLVNUXjkFOnhO67n+vDcQq7PFRBSDxC82vwnTSoerME6d57wC2hgd7TH2AhCRkoQ3gd6iVQNCfVvLfgnRT7rrXvG8tIlAzOeX1i07wpXO8uDFbDgjzPvTniB37+Fauxj32mJtW2Ta+YUP/V5Sg49Y/Q6Jd3KYTgFB0DjjY1+bZKkDp/WMhNV6RhytZyseAzHrOg5C5zmvgBaYLhvyZSVjBRBJZtKm2q3kCXqcJgZatRi5JbY87ZD8rl8dgC/MT3Jot8zKBUtOJvLeOXqTUWIxoRCrk6vrl70To0o7XG8JMYMDN8PEPGkjIoQZw+VVR9JR8863w/3oZNSkCWiLZs9uKAfT5S6koliYEgMhPWmluhwAbbcKX9Dj5L9oOOhXTFxSObGPk+V37LzH5MTvcIuMlRURzxNhiRBcQ+QfqOzmxCu9D436ry0tGiJU+syvNkh/z59z0O8QAYH90g84zzkJHm+oW8NyZr82nryXaCIPRkdPdm9esxEHofOcJ0C32JqrF/WJUO8IPAuB4fmYNcGrUNDYqHq2P9FBYn5g28QhbtghID6lcuHSv5lgwhIB9roREV0sEh4gvrfhD4gvKDlqyfxRwh/B9RvV9tavDkoqdQf8mPodBCXpAUEyefqMT9+4yHYQyDchyzozhromdoQjAGyp7kYAB6XznN8YOkr0AuixO44IhAyIIIZn2x1+58RtK612ICnUt4mJQndXTcacctujg0m8CVZYcteNJNlEHMyEOeB1I3uZV/kVfuC7JLVxdH7Hy+13mt6uPigvRgmPADGaSysCNWPeJEgd2GSw7KglbwbzGWLCGCecYsjpPHvJDst5Pgid57wCuntbXy9guvYI+YhAoo+a+EeyBxOd29f+QcT8zOl3xDs5FA2AL55x7JLKibvN0qYtPZQ/GJthZncwdxS+l8BkS4vBalGdX//HbRSm76NkSUwJXmIJ04CfuHTeddeVoonzeHJonYGf8K8ijq31LPZSsLl99/OP907WQUqugHZbd1tEgteGGHckBBGA+Wg753MkHwMKilblUBYkGG1YBUdFoVpmnevBbnk2qauqLT2dWJ0Ui9Bu9xcSmPuHE919tiEAQNjdcQvSqbj8DpYwbai8YkG69ozz7NTkhVZKQKuuWXh2lYh6n00G4+CE2eiol2kw75zng/TQWXDxS3viSgBDdSPjrQlODqlb960XHf4ldn8WW/QXYqf633awIuKLLF1W+de6ooYZ20uxEhN2Rqbn9v2OjeN3dLVc/UuT6n+SEiUcK1cQCNgDl1U5wlI+RNJtMigmcQEr/8iYyaDTedb3dG5+eFu+Oc8HB+hsfCzRlmHlJDgPjaNx7JKEqzNO5PYnQINCe1vKwNweX7+DWExIRPTuivrz3nCQLeEYaDHV9e+uA/EKpww09sSSBxHTOtDe9qqd2P3UYFthqyX9A7fGPum3tFLAD86t/Oebj8mPDt5qA4CE/Ry0nolhQm3C1I2YYouvnGT4xWHlJCIRAwEdUTKI2fl5NOt9FA3eZXTYE5PfQRBExEFSceIgFZYanKZf+WXMQbk7VKAY1ZYoNJmbxty5Gq0HC7e/8CPT2x2X30EwWnNZZUkwZ+77Jj45bGaApPLopYsJtJ96+2ilOp+0RD/v2f7ghOk85zHksJOiKfPinl5KhIirlMqGHRPNoWhx/I6HtzGM5XfEOVlzbTIgdKndPcalsERAm0ZdQ5KIr3IHlmPSJYl9EtGPDLRv+J8xyeyO39F946c7KZO6m+PzOwhRCAqCy49dsiThFsYEgdqGl+T7H3A38ePJtJGAjP7mwTmQSQO0rUVr6W8HpHPEaaEhUlDsL7Jf78zbi4iJboFoiXd6aAlLxN4JlSG/e3yLzWbo1VR+LrN/fMzrRlaqTKIbYpcNLb8D0euvfdsM9GXAzLHeL5MyVFp+QvdpV5wBookSSSegVZfPa5hFQo3uZDBOiMXGZP7YXVa2BlNwkDKesp0AwMA23gVBO9meOAIhARFA+Cv7bXPykNVafkfX1rJfio6eJIqt32EJS8q/fHylI0fiZ75uuNFLnIkN/9hVunU94pLZLb+DO7/44WclPfAIJUsJEuP9RAR+AJRWXA17GDUBY++0nksqLiIVzB6j3j5inJVVk93Umpl6jZT4XbAU0BYJzB9HlO7Y1mX5r0dUJvJglt9B0LciPr2DncLS8pKjzz4yuzDix5Etprzu3BNI1DnuUEHFmliYW7BpU3Zi44HM9WcxA72ryMQJbVwJLzUIBCXLZvzvr82fGB28VgNAka23S6z73KRYdGZXlErfNRSmTQtAD1U6zG+HKx02MSTgjeVHnDbbTWAewg67UHS0+x4xmVdj8jtcSzi/OuElncLSGfG8h5NEUJy8mpQfxCinCYhZTHq3VtGdLpOOv7ibmgxEqOIvz683/T3PUyIGv8NJHXBZRaU34yjXD/1gdPCc1nPt0lOJvLe6PpRjjZcm8kigv9//2kM7p+ogZXyAHqp06N9CNECiXB4kIDVLJapOzE9imA15LL8DRt+9ZxeBA0eEEAEoS1hqi3ebvLXV1Cw8uwqgy9y7jnWQool8EiM/6Hthwy405lyDFbSC2+9oSSGTuoO82P1ZSLSGJEsuR0ODd1BSB0Naz0EuWs9KdJjWRudF5znPIYfzkhl5Vky0G8TOY5MmYohSp+Y3MXT8jkhuEzPUn2Us/oOyJ3v81or6ZW/F0KnnAbcimyCY5EpWwXwxWo+ZGBGUmEzI0aCd2NZxTOxz19vkcMeOO01fdy+Ux2OW8IhYMoMGybK/nnX6le8EEcbZuJPQ2qpL65YdAaL3ioQxtJ6dzjOi9X3t6/44GfIEEwxoCCDU9+qGXYA86xJDk+2ERYLTR9aOJ9725nf4MZXxoUl5zIhLWDrDHSqoa52eDuJMrMA80Ln9oT+Me2Kz+h1f+fuXJDN4HyVLKJaXFzGUKCHUzLrKJofjGVubyAXkXUIcVMfUeibLeZZvxq7oFBigkeUkiJjHAc7G0SyiQeCTK448a2YeTgz3qZHC6Jtz0O9QYjTAdJENIw5EWLKnepVHL11MRKfFO1SwKqzDnVAPYmJb3W/s6Vsl6UGJ1Z9FoCQ9CPKSK8v+9ktzxqeDZ8lXQvhATvV2Ez3Vvblko022W/U0BHT2daK2EbElQYwhVrM9L3FqHuPobNhD1ejdIFHmLzH7s1iFJQ7maZ1cHuf5yPOvHYPEP6JUp9jo6Jneza8+dNA12FYLxl0tl/1KUgO/pkTp2FxwIkIUai6vnFlaXzcOHTyn9Vxb0cDsvyW2vBeIhPQN9n2bCkrul3MEFIj6nhTRryErMCFkQAyj/HP3qIjkJTlsUO3tbSkjcjsR59LcEsQHVFjai8Qf65KugBgCvco2TpqAhZzld6QGbwUriq3kIAIEpbYfei7JYTZE8fiDMeW9TFbn2evuyKvO82R4aAGauXvbY50QedyphZpsPZoI5wKLgn2u8E+oOX5HmL7T6LDHLqoxk0O2tWR1xsy6c0/AqApLe5H4x24gKYBiY8LXJD34wwmb2EZ7tT7847P3SF/3DvjB2BUTJ3WAZNkpM//vt08GkcRMDhmtrTo57+xaSEzylYgQKRDk27t3P96bU729MEOOjS6OlvUYWs/EYst3b6iorT85v2GH43e88uBLEBNXv4MAaFZeIuLEpQ7APFocaRtIxugZItCkFIkxd/a+8rPdTjf54CeWYPkdd7R0IZP+PiWS8SR4IZpLSxVV11haaWOM7LDBjkEQJN/Pyq+AO+seY0diMWEPDQ7cnt8iwKQB2ikcofdBY8J+kOtgJbbITkq9Z5Kem4j0KoiWeO+Q3UXofahrSO6psORI/HVl7wKrv46hDCQgUqKjtJHUreMu1Y1RwtO7d31b+nuz/A4ZOzlMgfySC6v+4avVVurggMmhux51bIJYXWlPfMcUsHEt2cy9nTuynOeW6Q5oGKCZu7Y+2k6Qxx2eDUhYRINB77GAyWfWa39315a1lt8Rqz+LIyyRf3wNSvciLDkSP3sfojg9Q+zdRYKYdX3tD/5pwmuwLS0GzcIdX7hmk0mnHnb8jjFPDiXMaCqrnKeOOGolgDFus9gdpabuhDNZeW+MmQwqkVAbmBswyeIxea1yDPX+NvqerIiSbTAfGWL/uEoqOd2+cGMeiSqNDMAIyarhZxjTqdvklQPXL3yR8+4tpmbh2bUgdb4r1fGY4BGNMJ+6E47fgYGuVdA65j1GAojAiZKrxw4HsuQrdS3AMna4ZHWeYfTGni1rnpxs8Zg8A9oeIYeRrDE67AEPNc60DDcKxslwy72EB915rzGZV8EqznGzcnp0S0vmLz8KaDFYfJ29CiMll7PyKoCxeoaIJvIYJnq6f3NvW95qsI7fsfuhhx6Q/t4/UVASg98BK5KeKHlXzae+/VdZJt++39jstJ7PWSBE54rJZDsF6/1+bINKCMkNeVvEUwdoCJqbeXD72peF5MHh3iiW4QbilSXzzzwqf5Jc2XS0kbu3PdZJxul3jA1oqz/HflXCg9Wfe3qeRl1DEkRXiXFNRcf4qyCCDPXaa8rf+12/UaHtjhTSqe+Q58cpUZKTOvDUrBmXj9xN9zRH4ufkR5VXWk7sKWL/gB/2SnzR4V+6OZw0nefxmje+OM+1exP9XUAuBoQBIhiJWPlVgVd66SDwZXt7oTVPj263zVBnbvNZ/T0AP1YdAQIh73Kg4VtAiy7H0iXEwfGiQzOkJro/kLFiozMvK/Vaa/4n1ul3bN92F8rKPw3lVUFHY1xycFIHQeKS2R9p/rddLe/uczuO7J2DGB1tMFHfkzBiRk0gRYiYQhilFRtFhD/jhQ3pfX9fYRkdxM8J5p9aUu3PfYbYO84mFgCRx2KiP3WZ352E9vb0CN+WB3Pb5zEr1zIFS8WEJqY6pkSSOr1v8/pfVi1YuY5VYqnojAYd6GdFEwcKUfoLnVvv/4zNEfJ85Lt6tUJTk5715fvv4Jlzr5SBngigMZyQGEqUcrTrpYs7Pv2+e7K/A4eJeeP8OStEuL11EAtWfAfE/wHjylkSGVLBCVV404putLfmd+Ld9qllFTwsi7lANbHvcajfW1rX8CKRI/GP5Z3tBdh+MfpW5+ny76VaXRra37NKyisvBzGPqf4vEDCLSpZfDeAePLe/2yzNnFssvEgKNRGcCA+djb9NzZHnHi2J0udAVO5yQ0PsK6Mzj3dvuf+0/Hpo9/x1DYlqVfUMUaz7f0KkyEj0J9H6F8ov+ZBNjA7wM4KIlO8Znbq7e8tPL51UuqQIgYhm/b+1j3F5zdslNaDHpHcSA0anpP2FN7/+pb/7C5qbGS2FD8apSgqzNuLUTv9w+NSOlJjQMPvvmLFgxdkYVVZ2IpNDp9+hcbuLNsZMnixDkE5g5X0oVkcnEgXRQkM12EnM8i2/w2Bw4FZLQzdj/20rdZDEnHnuZPQMxmFiB/mi2eJ6+psjGme6Eh7BgD+95/flw2y9NYwydxqT6cmhP0vMncPSJY3RT3RtWfvYyMRqUqzJ8Tv+9Id7TW/3K/ATMcj/IAkzID+4rO6qq5KWsCSu+8K0/UwGoFvcyeGDzwDmp8PEe+ulib1319QvPy/LwcjPjNvfPfjyhu0QuS8H8n8OIRfBWIlYmYKbzUP8Dsqkvh9Pv4OsDl5ZxfH9C88+y2qAtPKQs5men8kA9HBipnXmy2IiM6KWa700q8/ZmDPvR6XE2qyCFVScqJDAgBSLCbcGvR33YapuNmevaHXsuN0MxOV3iED5QGXN1SMTzEPdQ0/QxDc7HbgV95KXfK+YjLblM9HEvop0+oreLT+9M8+lLgJAVQsueJzZf5tIaGJ3bT1guJFQEqWau7be/7lJKdXtd4iF0UJm9lfW/JSqZ50vA71jlBlh17VIBszPTN+oWAz5Cc+LBj/2yt+e8yuIMIjMRJft9uMU9Odg9PmuFpyVOhBF9G8zZ678ye7di/qRt8J8g7Knd+YWEL0tF/L/frd6YjYm00uI7nCx89RVCrL6HX09t6iKqvNj8TtsX5yAvOCt0xjPoCAJnR6cYaeglfIccgzHsd3t634Hie4i9rPcA4ZoQ5yojyrkU+778pRxu1DA8jt2EJTKIZYe3TOQRxD9466t69ons9fe6Mmh43c8/qsHpL/n+fj9WQSSGTTT9hOmIkkNGj04ECuamEBwLXIyu9H1YtVCnSd21E3l/VN57cpFdstuzpt+h+V3yPetmjwdBKAtg9AYualQfBWu36iw4Rtpk8l8G/H1O1xZchp/iBgcyCQD2nrfrq3r2iH6y8Q+D7cONkLklXgKXx+RSOahlpvld6RvFxNmxtaXOFDs7MGI/kVv+5pfoQBECK257le7tt8pfT3d8fqzHAJGAPp2VUwyoLMxZjN39fd8VUz6zxiSZCUlJqNJJc6qqTv/by048hF62DJi/0sPPCdiHnb9WcYJRCLR+qbh+LwAzOl37P7Sx16RdP99lCyN2Z9lGpsV2IfJpIIpADQE2ETY1danEf4T0cgEkFhMZOD5/1l59JJj8kcvdfwO0Tdnj3hyz0KUEp1+sSfz6v0oEBHCYZ9hB1N3dqyS9KBxF4UPfROiqQA0nPdVvZvXrTVR+H3iQDkvSYARkKpiL3ELAEZjY77EHalL+h40JnyeWHFOyaFA3NWy27Dj6YFCESEcfr0mDRF0/scHn5CBvl/H7s8yrUMOgkgsenC+CPitAjRzKNEnxHaEHQ49dBSRl3h3df2K/4PWVo2GBm+i1zLQyGhvS4nR37Hl8NjJoYCZRUdd6WjwjuF3KTBz/A5Kp1a53pGHfhhNLFMIaBhgEw20r3vVmMz/ImICHKgISnRGk/Kbq2pXnGl1PCb6WNzWi70ocvodoANeMRq6aiQhkQcY/aPBlx5+ZSq7OR24hGfHUr/y7L3S1/Uy/MAJpYs5VD/sebEA7eXRjWigUfVsbV1dtWDlMlbJq0Sn3emWEACG5313xvwz39axvfXlCaZkCtCoOl5u3V614IL7WJVdKSY9dmGFoSAGYvRNyJtYzgS932pRHU3UM+u/fvx9NXPeJ6WvC2A1nXxuTNdomIMSaBJ/igE9XPVg+uU/iFFvJ/aG+MpitGblH6X9iruwePE5eHpT9oLqBO2fti4uJN8QE84XMWNcsRJD4vlGwme7t619unBKdfsxy+8g6dl9myktOw9GfKuHPR0SPMmeYsZK0o0JfY707hFJ8cEuk4Mxy3+YUb/sbYYTPwdEuVDHtmBTgSc6tapry/0ftupFbfqwCAon3vgQHTcaDmOnLobeJ/To2Lru10YynyD2FbLyXQRPdCYilbiupn7Fp4C2CIuv8/IwIDl8mqdZGUxGTrgcgh+TSx5Dk7bKGhoU2tqi6gUrbyeVvFp0OoLrLwHAEHkqMllWXoNniUZFK7B5nNq8ocAGgoBGxvyXgmr/iIdIee8UE+oRzDwBcahNf2Pvlg0/LYK6aOONuyZxhS0SbH9iMB32XGJMtI3IG3noAsAkFJXeXV2/pMGCucErTlHRChXQGLoutf2Rl0HhhQC6bJs22/AQYgwIpVClP66oXfaOIqiLVuCAHk4Su19c+7TW6UtIENquWmKFEkUbAmqUn1xT9NRFy9WmqBK/SYAGL9310F8S1cf9iUldPCLUZsAYIi4FeU3J6mOfTXU99CcL6nZTnLKiFSCgAQvOBi/d9eAfgsraneyVrIBk1T8tqEEcENRFfvWCrZmuh39XBHXRChjQI0Dd/bPfJMvrO8gLlg9dCsiCGvAUBxcGNce8mu586Ek0Nips2lScuaIVIqCHQZ3qfviJoHJBB6vEcohEI25iaIhoZv+CoOaYdPqX9/2iOG1FK2BAjww/Hn4iWXmskJc8i5iZWDGxUkRKgQnKKz07UbmwrmzuGx4Z3P18WJy+ou1tBVQ9aNNWhanlc9ULVu4mpY4ygnDoaFcAmMiA1AydCRcA+H0h9ZguWtH2Z1QcgqIdYuBpVAfsSNtWZOQVrWhFK1rRila0ohWtaEUrWtGKVrSiFa1oRSta0YpWtKIVrWhFK1rRila0ohWtaIVu/x9r9Ln6h64oDwAAAABJRU5ErkJggg=="
_LOGO_DATA_URI = f"data:image/png;base64,{_LOGO_PNG_B64}"
_page_icon = "💰"
if _logo_bytes:
    try:
        from PIL import Image
        _page_icon = Image.open(io.BytesIO(_logo_bytes))
    except Exception:
        pass

st.set_page_config(
    page_title="Finanzas WL",
    page_icon=_page_icon,
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ============================================================================
# PWA: meta tags inyectados al <head>
# ============================================================================
def inject_pwa_meta():
    icon_192_b64 = load_logo_b64(ICON_192_PATH)
    icon_512_b64 = load_logo_b64(ICON_512_PATH)

    if icon_192_b64 and icon_512_b64:
        icon_192_url = f"data:image/png;base64,{icon_192_b64}"
        icon_512_url = f"data:image/png;base64,{icon_512_b64}"
        apple_touch_icon = icon_192_url
        manifest_icons = (
            f'[{{"src":"{icon_192_url}","sizes":"192x192","type":"image/png","purpose":"any maskable"}},'
            f'{{"src":"{icon_512_url}","sizes":"512x512","type":"image/png","purpose":"any maskable"}}]'
        )
    else:
        # Fallback SVG con emoji (si el usuario aún no subió los íconos)
        svg_icon = (
            "data:image/svg+xml;utf8,"
            "<svg xmlns=%22http://www.w3.org/2000/svg%22 viewBox=%220 0 192 192%22>"
            "<rect width=%22192%22 height=%22192%22 fill=%22%23102250%22 rx=%2224%22/>"
            "<text x=%2296%22 y=%22130%22 font-size=%22120%22 text-anchor=%22middle%22 fill=%22white%22>💰</text>"
            "</svg>"
        )
        apple_touch_icon = svg_icon
        manifest_icons = f'[{{"src":"{svg_icon}","sizes":"192x192","type":"image/svg+xml"}}]'

    components.html(f"""
    <script>
    (function() {{
        try {{
            const parent = window.parent.document;
            const head = parent.head;

            const manifestStr = '{{"name":"Finanzas WL","short_name":"Finanzas","start_url":".","display":"standalone","background_color":"#ffffff","theme_color":"#102250","icons":{manifest_icons}}}';
            const manifestBlob = new Blob([manifestStr], {{type: 'application/manifest+json'}});
            const manifestURL = URL.createObjectURL(manifestBlob);

            const tags = [
                {{tag: 'meta', selector: 'meta[name="apple-mobile-web-app-capable"]', attrs: {{name: 'apple-mobile-web-app-capable', content: 'yes'}}}},
                {{tag: 'meta', selector: 'meta[name="mobile-web-app-capable"]', attrs: {{name: 'mobile-web-app-capable', content: 'yes'}}}},
                {{tag: 'meta', selector: 'meta[name="apple-mobile-web-app-status-bar-style"]', attrs: {{name: 'apple-mobile-web-app-status-bar-style', content: 'default'}}}},
                {{tag: 'meta', selector: 'meta[name="apple-mobile-web-app-title"]', attrs: {{name: 'apple-mobile-web-app-title', content: 'Finanzas WL'}}}},
                {{tag: 'meta', selector: 'meta[name="theme-color"]', attrs: {{name: 'theme-color', content: '#102250'}}}},
                {{tag: 'link', selector: 'link[rel="manifest"]', attrs: {{rel: 'manifest', href: manifestURL}}}},
                {{tag: 'link', selector: 'link[rel="apple-touch-icon"]', attrs: {{rel: 'apple-touch-icon', href: '{apple_touch_icon}'}}}},
            ];
            tags.forEach(t => {{
                if (!parent.querySelector(t.selector)) {{
                    const el = parent.createElement(t.tag);
                    Object.keys(t.attrs).forEach(k => el.setAttribute(k, t.attrs[k]));
                    head.appendChild(el);
                }}
            }});
        }} catch(e) {{
            console.warn('PWA inject failed:', e);
        }}
    }})();
    </script>
    """, height=0)

inject_pwa_meta()

# ============================================================================
# CSS BRANDING
# ============================================================================
CSS_BRANDING = f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Open+Sans:wght@400;500;600;700&family=Poppins:wght@500;600;700&display=swap');

:root {{
    --navy: {NAVY};
    --cyan: {CYAN};
    --green: {GREEN};
    --orange: {ORANGE};
    --red: {RED};
    --grey: {GREY};
    --navy-hover: {NAVY_HOVER};
    --navy-light: {NAVY_LIGHT};
    --bg-soft: {BG_SOFT};
    --border: {BORDER};
    --text-muted: {TEXT_MUTED};
}}

html, body, [class*="css"], .stMarkdown, .stTextInput, .stSelectbox, .stNumberInput,
.stDateInput, .stRadio, .stButton, .stMetric, .stDataFrame, .stTabs, .stForm {{
    font-family: 'Open Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
}}
h1, h2, h3, h4, h5, h6 {{
    font-family: 'Poppins', 'Open Sans', sans-serif !important;
    color: var(--navy) !important;
    font-weight: 600 !important;
    letter-spacing: -0.01em;
}}
h1 {{ font-size: 1.85rem !important; }}
h2 {{ font-size: 1.4rem !important; }}
h3 {{ font-size: 1.15rem !important; }}

.block-container {{
    padding-top: 1.5rem !important;
    padding-bottom: 3rem !important;
    max-width: 720px !important;
}}

section[data-testid="stSidebar"] {{
    background-color: white !important;
    border-right: 1px solid var(--border);
}}
section[data-testid="stSidebar"] .stRadio > div {{ gap: 0.25rem; }}
section[data-testid="stSidebar"] .stRadio label {{
    padding: 0.6rem 0.75rem;
    border-radius: 8px;
    transition: background 0.15s;
}}
section[data-testid="stSidebar"] .stRadio label:hover {{
    background: var(--navy-light);
}}

.stFormSubmitButton button,
div[data-testid="stFormSubmitButton"] button {{
    background-color: var(--navy) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    padding: 0.55rem 1.25rem !important;
    transition: background 0.15s ease, transform 0.05s ease !important;
    box-shadow: 0 1px 2px rgba(16, 34, 80, 0.15) !important;
}}
.stFormSubmitButton button:hover,
div[data-testid="stFormSubmitButton"] button:hover {{
    background-color: var(--navy-hover) !important;
    color: white !important;
    box-shadow: 0 2px 4px rgba(16, 34, 80, 0.2) !important;
}}

.stButton > button {{
    border-radius: 8px !important;
    border: 1px solid var(--border) !important;
    color: var(--navy) !important;
    background: white !important;
    font-weight: 500 !important;
    transition: all 0.15s ease !important;
}}
.stButton > button:hover {{
    border-color: var(--cyan) !important;
    color: var(--cyan) !important;
    background: white !important;
}}
.stButton > button:focus {{
    box-shadow: 0 0 0 3px rgba(21, 149, 188, 0.2) !important;
    border-color: var(--cyan) !important;
}}

.stTextInput input, .stNumberInput input, .stDateInput input,
.stTextArea textarea {{
    border-radius: 8px !important;
    border: 1px solid var(--border) !important;
    font-family: 'Open Sans', sans-serif !important;
    transition: border-color 0.15s, box-shadow 0.15s !important;
}}
.stTextInput input:focus, .stNumberInput input:focus, .stDateInput input:focus,
.stTextArea textarea:focus {{
    border-color: var(--cyan) !important;
    box-shadow: 0 0 0 3px rgba(21, 149, 188, 0.15) !important;
}}
.stSelectbox > div > div {{
    border-radius: 8px !important;
    border: 1px solid var(--border) !important;
}}
.stSelectbox > div > div:focus-within {{
    border-color: var(--cyan) !important;
    box-shadow: 0 0 0 3px rgba(21, 149, 188, 0.15) !important;
}}

.stTabs [data-baseweb="tab-list"] {{
    gap: 1.5rem;
    border-bottom: 1px solid var(--border);
}}
.stTabs [data-baseweb="tab"] {{
    height: auto;
    padding: 0.6rem 0.25rem !important;
    background-color: transparent !important;
    color: var(--text-muted) !important;
    font-weight: 500 !important;
}}
.stTabs [aria-selected="true"] {{
    color: var(--navy) !important;
    border-bottom: 2.5px solid var(--cyan) !important;
}}

.stDataFrame [data-testid="stTable"] thead tr th,
.stDataFrame thead tr th {{
    background-color: var(--bg-soft) !important;
    color: var(--navy) !important;
    font-weight: 600 !important;
    border-bottom: 2px solid var(--navy-light) !important;
}}

.metric-card {{
    background: white;
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 14px 16px;
    box-shadow: 0 1px 2px rgba(16, 34, 80, 0.04);
    transition: box-shadow 0.15s ease;
    min-height: 94px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    container-type: inline-size;
}}
.metric-card:hover {{
    box-shadow: 0 4px 12px rgba(16, 34, 80, 0.08);
}}
.metric-card .metric-label {{
    font-size: 0.78rem;
    color: var(--text-muted);
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    margin-bottom: 6px;
}}
.metric-card .metric-value {{
    font-family: 'Poppins', sans-serif;
    font-size: clamp(0.95rem, 9cqw, 1.45rem);
    font-weight: 600;
    color: var(--navy);
    line-height: 1.1;
    white-space: nowrap;
    font-variant-numeric: tabular-nums;
}}
.metric-card .metric-sublabel {{
    font-size: 0.68rem;
    color: var(--text-muted);
    font-weight: 500;
    margin-top: 3px;
    line-height: 1.1;
}}

/* Ocultar los anchors ("clips") que Streamlit agrega a los títulos */
[data-testid="stHeaderActionElements"] {{ display: none !important; }}
.stMarkdown h1 a, .stMarkdown h2 a, .stMarkdown h3 a,
.stMarkdown h4 a, .stMarkdown h5 a, .stMarkdown h6 a {{ display: none !important; }}

/* Mobile */
@media (max-width: 640px) {{
    .block-container {{ padding-left: 0.9rem !important; padding-right: 0.9rem !important; }}
    .metric-card {{ min-height: 80px; padding: 12px 14px; }}
    .page-header .page-title {{ font-size: 1.4rem !important; }}
}}
.metric-card.metric-green .metric-value {{ color: var(--green); }}
.metric-card.metric-orange .metric-value {{ color: var(--orange); }}
.metric-card.metric-cyan .metric-value {{ color: var(--cyan); }}
.metric-card.metric-red .metric-value {{ color: var(--red); }}

/* Toolbar de Streamlit (Share / GitHub / lápiz) oculta */
[data-testid="stToolbar"] {{ display: none !important; }}

/* El header de Streamlit es una franja transparente que flota ENCIMA del
   contenido y bloquea los taps sobre la primera fila del menú en mobile.
   Ya no lo necesitamos (la navegación vive en el cuerpo): fuera. */
header[data-testid="stHeader"] {{
    display: none !important;
}}

/* Sidebar eliminada: la navegación y la sesión viven en el cuerpo */
section[data-testid="stSidebar"] {{ display: none !important; }}
div[data-testid="collapsedControl"] {{ display: none !important; }}

/* Los componentes invisibles (cookies) dejan un hueco arriba: ocultarlo.
   Los iframes siguen ejecutándose aunque no se muestren. */
.element-container:has(iframe) {{ display: none !important; }}
div[data-testid="stIFrame"] {{ display: none !important; }}
.stApp .block-container {{ padding-top: 1.2rem !important; }}

/* Navegación horizontal: chips prolijos */
div[role="radiogroup"][aria-label="Menú"] {{ gap: 4px; flex-wrap: wrap; }}

.google-btn {{
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 10px;
    width: 100%;
    padding: 0.55rem 1rem;
    background: white;
    border: 1px solid var(--border);
    border-radius: 8px;
    color: var(--navy) !important;
    font-weight: 600;
    font-size: 0.95rem;
    text-decoration: none !important;
    box-shadow: 0 1px 2px rgba(16, 34, 80, 0.05);
    transition: box-shadow 0.15s ease;
}}
.google-btn:hover {{
    box-shadow: 0 3px 8px rgba(16, 34, 80, 0.12);
}}

.mov-row {{
    background: white;
    border: 1px solid var(--border);
    border-left: 3px solid var(--grey);
    border-radius: 8px;
    padding: 10px 14px;
    margin-bottom: 6px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 12px;
}}
.mov-row.mov-ingreso {{ border-left-color: var(--green); }}
.mov-row.mov-fijo    {{ border-left-color: var(--navy); }}
.mov-row.mov-variable{{ border-left-color: var(--orange); }}
.mov-row.mov-ahorro  {{ border-left-color: var(--cyan); }}
.mov-row .mov-info {{ flex: 1; min-width: 0; }}
.mov-row .mov-cat {{
    font-weight: 600;
    color: var(--navy);
    font-size: 0.95rem;
}}
.mov-row .mov-desc {{
    color: var(--text-muted);
    font-size: 0.82rem;
    margin-top: 2px;
}}
.mov-row .mov-monto {{
    font-family: 'Poppins', sans-serif;
    font-weight: 600;
    color: var(--navy);
    white-space: nowrap;
}}
.mov-row.mov-ingreso .mov-monto {{ color: var(--green); }}

.login-footer {{
    text-align: center;
    margin-top: 24px;
    color: var(--text-muted);
    font-size: 0.75rem;
}}

.page-header {{ margin-bottom: 1rem; }}
.page-header .page-title {{
    color: var(--navy);
    font-family: 'Poppins', sans-serif;
    font-weight: 600;
    font-size: 1.5rem;
    margin: 0;
}}
.page-header .page-caption {{
    color: var(--text-muted);
    font-size: 0.85rem;
    margin-top: 2px;
}}

hr {{
    border-color: var(--border) !important;
    margin: 1.25rem 0 !important;
}}

.stAlert {{ border-radius: 8px !important; }}

footer {{ visibility: hidden !important; }}
#MainMenu {{ visibility: hidden !important; }}
.modebar {{ display: none !important; }}
</style>
"""

st.markdown(CSS_BRANDING, unsafe_allow_html=True)

# ============================================================================
# CONEXIÓN A SUPABASE
# ============================================================================
def init_supabase() -> Client:
    # SIN @st.cache_resource: cache_resource es GLOBAL al servidor (compartido
    # entre todos los usuarios conectados). El cliente se crea por ejecución y
    # la sesión de cada usuario se rehidrata desde st.session_state, que sí es
    # individual por navegador.
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

sb = init_supabase()

# ============================================================================
# ESTADO DE SESIÓN
# ============================================================================
if "user" not in st.session_state:
    st.session_state.user = None
if "sb_tokens" not in st.session_state:
    st.session_state.sb_tokens = None

# ----------------------------------------------------------------------------
# COOKIE DE SESIÓN PERSISTENTE
# Guarda el refresh token 30 días para que cerrar el navegador no desloguee.
# ----------------------------------------------------------------------------
cookie_manager = stx.CookieManager(key="wl_cookies")

def _guardar_cookie_sesion(refresh_token: str, key: str):
    try:
        cookie_manager.set(
            "wl_rt", refresh_token,
            expires_at=datetime.now() + timedelta(days=30),
            key=key,
        )
    except Exception:
        pass

# Si no hay sesión en esta pestaña pero hay cookie, restaurar.
if not st.session_state.sb_tokens and not st.session_state.user:
    _rt_cookie = None
    try:
        _rt_cookie = cookie_manager.get("wl_rt")
    except Exception:
        pass
    if _rt_cookie:
        try:
            r = sb.auth.refresh_session(_rt_cookie)
            if r and r.session:
                st.session_state.sb_tokens = {
                    "access_token": r.session.access_token,
                    "refresh_token": r.session.refresh_token,
                }
                st.session_state.user = r.user
                # Supabase rota el refresh token: guardar el nuevo.
                _guardar_cookie_sesion(r.session.refresh_token, "ck_restore")
        except Exception:
            try:
                cookie_manager.delete("wl_rt", key="ck_del_invalid")
            except Exception:
                pass

# Rehidratar la sesión de ESTE usuario en cada ejecución del script.
# Si el access token venció, set_session lo refresca solo con el refresh token;
# en ese caso guardamos los tokens nuevos.
if st.session_state.sb_tokens:
    try:
        sb.auth.set_session(
            st.session_state.sb_tokens["access_token"],
            st.session_state.sb_tokens["refresh_token"],
        )
        s = sb.auth.get_session()
        if s:
            if s.refresh_token != st.session_state.sb_tokens.get("refresh_token"):
                # Hubo refresh (rotación): actualizar también la cookie.
                _guardar_cookie_sesion(s.refresh_token, "ck_rotate")
            st.session_state.sb_tokens = {
                "access_token": s.access_token,
                "refresh_token": s.refresh_token,
            }
    except Exception:
        # Token inválido o vencido sin posibilidad de refresh: volver al login.
        st.session_state.user = None
        st.session_state.sb_tokens = None

# ============================================================================
# AUTENTICACIÓN
# ============================================================================
def do_signup(email, password):
    try:
        r = sb.auth.sign_up({"email": email, "password": password})
        return r.user, None
    except Exception as e:
        return None, str(e)

def do_signin(email, password):
    try:
        r = sb.auth.sign_in_with_password({"email": email, "password": password})
        if r.user and r.session:
            # Guardar los tokens en la sesión de ESTE navegador.
            st.session_state.sb_tokens = {
                "access_token": r.session.access_token,
                "refresh_token": r.session.refresh_token,
            }
            _guardar_cookie_sesion(r.session.refresh_token, "ck_login")
            # Si es la primera vez que entra (todavía no tiene categorías),
            # le sembramos el set por defecto.
            try:
                seed_categorias_si_vacio(r.user.id)
            except Exception:
                pass
        return r.user, None
    except Exception as e:
        return None, str(e)

def do_signout():
    try:
        sb.auth.sign_out()
    except Exception:
        pass
    try:
        cookie_manager.delete("wl_rt", key="ck_del_logout")
    except Exception:
        pass
    st.session_state.user = None
    st.session_state.sb_tokens = None
    st.rerun()

# ============================================================================
# ACCESO A DATOS: PLAN / SUSCRIPCIÓN
# ============================================================================
@st.cache_data(ttl=30, show_spinner=False)
def _plan_usuario(user_id: str) -> dict:
    """Lee la suscripción del usuario. Sin fila (o si falla la consulta) => gratis.
    A prueba de fallos: nunca rompe la app ni regala Premium por error, así que
    el .py se puede pegar aun antes de crear la tabla en Supabase."""
    try:
        res = (
            sb.table("suscripciones").select("plan, vence")
            .eq("user_id", user_id).limit(1).execute().data
        )
        if not res:
            return {"plan": "gratis", "vence": None}
        fila = res[0]
        return {
            "plan": (fila.get("plan") or "gratis").strip().lower(),
            "vence": fila.get("vence"),
        }
    except Exception:
        return {"plan": "gratis", "vence": None}

def es_premium(user) -> bool:
    """True solo si el usuario tiene plan 'premium' vigente (sin vencimiento, o
    con vencimiento en el futuro)."""
    if user is None:
        return False
    info = _plan_usuario(user.id)
    if info.get("plan") != "premium":
        return False
    vence = info.get("vence")
    if not vence:
        return True
    try:
        return pd.to_datetime(vence).date() >= date.today()
    except Exception:
        return True

def boton_premium(label: str, key: str):
    """Botón visible con candado para una función Premium. Al tocarlo promociona
    el plan en vez de ejecutar la acción."""
    if st.button(f"🔒 {label}", key=key):
        st.info(MSG_PREMIUM)


# ============================================================================
# ACCESO A DATOS: CATEGORÍAS
# ============================================================================
CATEGORIAS_DEFAULT = {
    "Ingreso": [
        "Sueldo",
        "Honorarios",
        "Alquileres cobrados",
        "Rendimientos / Intereses",
        "Otros ingresos",
    ],
    "Gasto Fijo": [
        "Alquiler / Expensas",
        "Servicios (luz/gas/agua)",
        "Internet",
        "Telefonía",
        "Prepaga / Obra social",
        "Seguros",
        "Cuotas crédito",
        "Tarjeta de crédito",
    ],
    "Gasto Variable": [
        "Supermercado",
        "Comida / Restaurantes",
        "Transporte / Combustible",
        "Salud (farmacia/médicos)",
        "Indumentaria",
        "Educación",
        "Entretenimiento",
        "Hogar y mantenimiento",
        "Regalos",
        "Viajes",
    ],
    "Ahorro": [
        "Plazo fijo",
        "Dólares",
        "Inversiones (CEDEARs / FCI / Acciones)",
        "Caja de ahorro",
        "Otros",
    ],
}

def seed_categorias_si_vacio(user_id: str) -> bool:
    """Si el usuario no tiene categorías, crea el set por defecto.
    Devuelve True si sembró, False si ya tenía."""
    try:
        res = (
            sb.table("categorias").select("id", count="exact")
            .eq("user_id", user_id).limit(1).execute()
        )
        if (res.count or 0) > 0:
            return False
        rows = []
        for tipo, nombres in CATEGORIAS_DEFAULT.items():
            for nombre in nombres:
                rows.append({
                    "user_id": user_id,
                    "nombre": nombre,
                    "tipo": tipo,
                    "activa": True,
                })
        if rows:
            sb.table("categorias").insert(rows).execute()
        return True
    except Exception:
        # No romper el login si falla el seed
        return False
@st.cache_data(ttl=30, show_spinner=False)
def get_categorias(user_id: str, tipo: str = None, solo_activas: bool = True):
    q = sb.table("categorias").select("*").eq("user_id", user_id)
    if solo_activas:
        q = q.eq("activa", True)
    if tipo:
        q = q.eq("tipo", tipo)
    return q.order("nombre").execute().data

def categoria_existe(user_id, nombre, tipo, excluir_id=None):
    q = (
        sb.table("categorias").select("id")
        .eq("user_id", user_id).eq("tipo", tipo).eq("nombre", nombre)
    )
    res = q.execute().data
    if excluir_id:
        res = [r for r in res if r["id"] != excluir_id]
    return len(res) > 0

def insert_categoria(user_id, nombre, tipo):
    payload = {"user_id": user_id, "nombre": nombre, "tipo": tipo, "activa": True}
    return sb.table("categorias").insert(payload).execute()

def update_categoria(user_id, cat_id, nombre=None, activa=None):
    payload = {}
    if nombre is not None:
        payload["nombre"] = nombre
    if activa is not None:
        payload["activa"] = activa
    if not payload:
        return None
    return (
        sb.table("categorias").update(payload)
        .eq("id", cat_id).eq("user_id", user_id).execute()
    )

def renombrar_categoria_en_movimientos(user_id, nombre_viejo, nombre_nuevo):
    return (
        sb.table("movimientos").update({"categoria": nombre_nuevo})
        .eq("user_id", user_id).eq("categoria", nombre_viejo).execute()
    )

def contar_movimientos_categoria(user_id, nombre, tipo):
    res = (
        sb.table("movimientos").select("id", count="exact")
        .eq("user_id", user_id).eq("categoria", nombre).eq("tipo", tipo).execute()
    )
    return res.count or 0

@st.cache_data(ttl=30, show_spinner=False)
def conteo_movs_por_categoria(user_id: str, tipo: str) -> dict:
    """Una sola consulta: {nombre_categoria: cantidad de movimientos} para un tipo."""
    res = (
        sb.table("movimientos").select("categoria")
        .eq("user_id", user_id).eq("tipo", tipo).execute().data
    )
    out = {}
    for r in res:
        out[r["categoria"]] = out.get(r["categoria"], 0) + 1
    return out

def contar_categorias_propias(user_id: str) -> int:
    """Cantidad de categorías creadas por el usuario (las que NO pertenecen al set
    por defecto), sumando todos los tipos. Sirve para el límite del plan gratis."""
    todas = get_categorias(user_id, solo_activas=False)  # todos los tipos
    return sum(
        1 for c in todas
        if c["nombre"] not in CATEGORIAS_DEFAULT.get(c["tipo"], [])
    )

# ============================================================================
# ACCESO A DATOS: MOVIMIENTOS
# ============================================================================
@st.cache_data(ttl=30, show_spinner=False)
def get_movimientos(user_id: str, limit: int = None):
    q = (
        sb.table("movimientos")
        .select("*")
        .eq("user_id", user_id)
        .order("fecha_devengo", desc=True)
        .order("id", desc=True)
    )
    if limit:
        q = q.limit(limit)
    return q.execute().data

def insert_movimiento(user_id, fecha_devengo, tipo, categoria, concepto, monto, cuotas, inicio_pago):
    payload = {
        "user_id": user_id,
        "fecha_devengo": fecha_devengo.isoformat(),
        "tipo": tipo,
        "categoria": categoria,
        "concepto": concepto or None,
        "monto_total": float(monto),
        "cuotas": int(cuotas),
        "inicio_pago": inicio_pago.isoformat(),
    }
    return sb.table("movimientos").insert(payload).execute()

def update_movimiento(user_id, mov_id, fecha_devengo, tipo, categoria, concepto, monto, cuotas, inicio_pago):
    payload = {
        "fecha_devengo": fecha_devengo.isoformat(),
        "tipo": tipo,
        "categoria": categoria,
        "concepto": concepto or None,
        "monto_total": float(monto),
        "cuotas": int(cuotas),
        "inicio_pago": inicio_pago.isoformat(),
    }
    return (
        sb.table("movimientos").update(payload)
        .eq("id", mov_id).eq("user_id", user_id).execute()
    )

def delete_movimiento(user_id, mov_id):
    return (
        sb.table("movimientos").delete()
        .eq("id", mov_id).eq("user_id", user_id).execute()
    )

# ============================================================================
# ACCESO A DATOS: SALDOS INICIALES MANUALES
# ============================================================================
@st.cache_data(ttl=30, show_spinner=False)
def get_saldos_iniciales_manuales(user_id: str) -> dict:
    """
    Retorna un dict {date(yyyy,mm,01): ajuste_float} con los ajustes manuales
    cargados por el usuario.

    SEMÁNTICA (v3.4):
    El valor guardado es un DELTA (ajuste aditivo) sobre el saldo inicial calculado
    automáticamente del mes anterior. Para el primer mes con datos, el saldo final
    del mes anterior es 0, entonces el ajuste funciona como saldo de partida absoluto.

    Fórmula completa: Saldo inicial M = Saldo final M-1 + ajustes.get(M, 0)
    """
    res = (
        sb.table("saldos_iniciales").select("mes, monto")
        .eq("user_id", user_id).execute().data
    )
    out = {}
    for r in res:
        d = pd.to_datetime(r["mes"]).date().replace(day=1)
        out[d] = float(r["monto"])
    return out

def upsert_saldo_inicial(user_id, mes_date, monto):
    """
    Inserta o actualiza el saldo inicial manual del mes indicado.
    `mes_date` debe ser un objeto date; se normaliza al primer día del mes.
    """
    mes_norm = mes_date.replace(day=1)
    payload = {
        "user_id": user_id,
        "mes": mes_norm.isoformat(),
        "monto": float(monto),
        "updated_at": "now()",
    }
    # Si ya existe (user_id+mes), update; si no, insert.
    existing = (
        sb.table("saldos_iniciales").select("id")
        .eq("user_id", user_id).eq("mes", mes_norm.isoformat()).execute().data
    )
    if existing:
        return (
            sb.table("saldos_iniciales")
            .update({"monto": float(monto), "updated_at": "now()"})
            .eq("user_id", user_id).eq("mes", mes_norm.isoformat()).execute()
        )
    return sb.table("saldos_iniciales").insert(payload).execute()

def delete_saldo_inicial(user_id, mes_date):
    mes_norm = mes_date.replace(day=1)
    return (
        sb.table("saldos_iniciales").delete()
        .eq("user_id", user_id).eq("mes", mes_norm.isoformat()).execute()
    )

# ============================================================================
# HELPERS
# ============================================================================
def fmt_money(n, decimals: int = 2):
    """
    Formato monetario es-AR: $77.569,90 / -$5.500,25
    - Punto como separador de miles
    - Coma como separador decimal
    - Signo negativo antes del símbolo $
    - Por defecto 2 decimales; pasar decimals=0 para enteros redondeados.
    """
    try:
        if n is None or (isinstance(n, float) and pd.isna(n)):
            return "—"
        v = float(n)
        signo = "-" if v < 0 else ""
        # Truco de intercambio: formato US primero, luego swap.
        s = f"{abs(v):,.{decimals}f}"             # '77,569.90'
        s = s.replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{signo}${s}"
    except Exception:
        return str(n)

def df_movimientos(user_id):
    movs = get_movimientos(user_id)
    if not movs:
        return pd.DataFrame()
    df = pd.DataFrame(movs)
    df["fecha_devengo"] = pd.to_datetime(df["fecha_devengo"])
    df["inicio_pago"] = pd.to_datetime(df["inicio_pago"])
    df["monto_total"] = df["monto_total"].astype(float)
    df["concepto"] = df["concepto"].fillna("")
    return df

def expandir_a_caja(df):
    if df.empty:
        return df
    filas = []
    for _, m in df.iterrows():
        cuotas = int(m["cuotas"])
        monto_cuota = float(m["monto_total"]) / cuotas
        for i in range(cuotas):
            mes_pago = m["inicio_pago"] + relativedelta(months=i)
            filas.append({
                "tipo": m["tipo"],
                "categoria": m["categoria"],
                "concepto": m["concepto"],
                "monto_cuota": monto_cuota,
                "mes_pago": mes_pago.replace(day=1),
                "n_cuota": i + 1,
                "total_cuotas": cuotas,
            })
    return pd.DataFrame(filas)

# Estilos de marca para los Excel descargables
_XL_FONT = "Arial"
_XL_MONEY = '[$-2C0A]"$ "#,##0.00;[Red][$-2C0A]"-$ "#,##0.00'
_XL_PCT = '[$-2C0A]0.0%'
_XL_NAVY = "102250"; _XL_GREY = "6C6D6D"; _XL_WHITE = "FFFFFF"
_XL_DARK = "22263A"; _XL_NAVYSOFT = "E8ECF5"; _XL_ZEBRA = "F3F5FA"
_XL_NRULE = Side(style="thin", color=_XL_NAVY)


def _xl_setup_sheet(ws, report_name, periodo, ncols):
    """Encabezado de marca + ajuste de página + sin grilla."""
    try:
        img = _XLImage(io.BytesIO(base64.b64decode(_LOGO_PNG_B64)))
        img.width = 52; img.height = 52
        ws.add_image(img, "A1")
    except Exception:
        pass
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    last = _col(ncols)
    for r in (1, 2, 3):
        ws.merge_cells(f"B{r}:{last}{r}")
    ws["B1"] = "WL HNOS & ASOC"
    ws["B1"].font = Font(name=_XL_FONT, size=14, bold=True, color=_XL_NAVY)
    ws["B1"].alignment = Alignment(vertical="center")
    ws["B2"] = report_name
    ws["B2"].font = Font(name=_XL_FONT, size=11, bold=True, color=_XL_DARK)
    ws["B3"] = f"Período: {periodo}   ·   Emitido: {date.today().strftime('%d/%m/%Y')}"
    ws["B3"].font = Font(name=_XL_FONT, size=9, color=_XL_GREY)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 17
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 6
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 9
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5


def _xl_table_header(ws, row, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = Font(name=_XL_FONT, size=10, bold=True, color=_XL_WHITE)
        c.fill = PatternFill("solid", fgColor=_XL_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=_XL_NRULE)
    ws.row_dimensions[row].height = 24


def _xl_money(ws, row, col, val, bold=False, fill=None, top=False):
    c = ws.cell(row=row, column=col, value=float(val))
    c.number_format = _XL_MONEY
    c.font = Font(name=_XL_FONT, size=10, bold=bold, color=(_XL_NAVY if bold else _XL_DARK))
    c.alignment = Alignment(horizontal="right")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if top:
        c.border = Border(top=_XL_NRULE)
    return c


def _xl_text(ws, row, col, val, bold=False, color=None, fill=None,
             align="left", size=10, italic=False, top=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=_XL_FONT, size=size, bold=bold, italic=italic,
                  color=(color or _XL_DARK))
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if top:
        c.border = Border(top=_XL_NRULE)
    return c


def _xl_footer(ws, row, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1,
                value=("WL HNOS & ASOC   ·   wlhnos.vercel.app   ·   "
                       f"Reporte generado el {date.today().strftime('%d/%m/%Y')}"))
    c.font = Font(name=_XL_FONT, size=8, italic=True, color=_XL_GREY)
    c.alignment = Alignment(horizontal="left")


_ORDEN_TIPOS = ["Ingreso", "Gasto Fijo", "Gasto Variable", "Ahorro"]


def build_xlsx_movimientos(df, periodo):
    """Workbook de marca con dos hojas: Resumen por categoría + Detalle.
    df: columnas fecha_devengo, inicio_pago, tipo, categoria, concepto,
    monto_total, cuotas. Montos como números reales."""
    wb = Workbook()

    # ---------- Hoja 1: Resumen por categoría ----------
    ws = wb.active; ws.title = "Resumen por categoría"
    _xl_setup_sheet(ws, "Movimientos por categoría · Resumen", periodo, 4)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    HR = 5
    _xl_table_header(ws, HR, ["Tipo", "Categoría", "Total", "% del tipo"])
    ws.freeze_panes = f"A{HR + 1}"
    resumen = df.groupby(["tipo", "categoria"], as_index=False)["monto_total"].sum()
    r = HR + 1
    for tipo in _ORDEN_TIPOS:
        grp = resumen[resumen["tipo"] == tipo].sort_values("monto_total", ascending=False)
        if grp.empty:
            continue
        sub = float(grp["monto_total"].sum())
        for _, row_ in grp.iterrows():
            _xl_text(ws, r, 1, tipo, color=_XL_GREY)
            _xl_text(ws, r, 2, row_["categoria"])
            _xl_money(ws, r, 3, float(row_["monto_total"]))
            pc = ws.cell(row=r, column=4,
                         value=(float(row_["monto_total"]) / sub if sub else 0))
            pc.number_format = _XL_PCT
            pc.font = Font(name=_XL_FONT, size=10, color=_XL_DARK)
            pc.alignment = Alignment(horizontal="right")
            r += 1
        ws.merge_cells(f"A{r}:B{r}")
        _xl_text(ws, r, 1, f"Subtotal {tipo}", bold=True, color=_XL_NAVY,
                 fill=_XL_NAVYSOFT, top=True)
        _xl_text(ws, r, 2, None, fill=_XL_NAVYSOFT, top=True)
        _xl_money(ws, r, 3, sub, bold=True, fill=_XL_NAVYSOFT, top=True)
        d = ws.cell(row=r, column=4, value=1.0); d.number_format = _XL_PCT
        d.font = Font(name=_XL_FONT, size=10, bold=True, color=_XL_NAVY)
        d.alignment = Alignment(horizontal="right")
        d.fill = PatternFill("solid", fgColor=_XL_NAVYSOFT)
        d.border = Border(top=_XL_NRULE)
        r += 1
    _xl_footer(ws, r + 1, 4)

    # ---------- Hoja 2: Detalle de movimientos ----------
    ws2 = wb.create_sheet("Detalle de movimientos")
    _xl_setup_sheet(ws2, "Movimientos por categoría · Detalle", periodo, 7)
    for i, w in enumerate([12, 14, 20, 26, 16, 8, 13]):
        ws2.column_dimensions[_col(i + 1)].width = w
    HR2 = 5
    _xl_table_header(ws2, HR2, ["Fecha", "Tipo", "Categoría", "Concepto",
                                "Monto", "Cuotas", "Inicio pago"])
    ws2.freeze_panes = f"A{HR2 + 1}"
    dd = df.copy()
    dd["__ord"] = dd["tipo"].map({t: i for i, t in enumerate(_ORDEN_TIPOS)}).fillna(99)
    dd = dd.sort_values(["__ord", "categoria", "fecha_devengo"])
    r = HR2 + 1
    z = 0
    for _, m in dd.iterrows():
        fill = _XL_ZEBRA if z % 2 == 1 else None
        _xl_text(ws2, r, 1, pd.to_datetime(m["fecha_devengo"]).strftime("%d/%m/%Y"), fill=fill)
        _xl_text(ws2, r, 2, m["tipo"], color=_XL_GREY, fill=fill)
        _xl_text(ws2, r, 3, m["categoria"], fill=fill)
        _xl_text(ws2, r, 4, (m["concepto"] or ""), fill=fill)
        _xl_money(ws2, r, 5, float(m["monto_total"]), fill=fill)
        _xl_text(ws2, r, 6, int(m["cuotas"]), align="center", fill=fill)
        _xl_text(ws2, r, 7, pd.to_datetime(m["inicio_pago"]).strftime("%d/%m/%Y"),
                 align="center", fill=fill)
        r += 1; z += 1
    # Totales separados por tipo
    r += 1
    _xl_text(ws2, r, 1, "Totales por tipo", bold=True, color=_XL_NAVY)
    r += 1
    primero = True
    for tipo in _ORDEN_TIPOS:
        sub_df = dd[dd["tipo"] == tipo]
        if sub_df.empty:
            continue
        ws2.merge_cells(f"A{r}:D{r}")
        _xl_text(ws2, r, 1, f"Total {tipo}", bold=True, color=_XL_DARK, top=primero)
        for cc in (2, 3, 4):
            ws2.cell(row=r, column=cc).border = Border(top=_XL_NRULE) if primero else Border()
        _xl_money(ws2, r, 5, float(sub_df["monto_total"].sum()), bold=True, top=primero)
        primero = False
        r += 1
    _xl_footer(ws2, r + 1, 7)

    out = io.BytesIO(); wb.save(out); return out.getvalue()


def build_xlsx_flujo(pivot_final, periodo):
    """Workbook de marca con el Flujo de Fondos (matriz concepto x mes).
    pivot_final: index = conceptos, columns = etiquetas de mes (str), floats."""
    wb = Workbook(); ws = wb.active; ws.title = "Flujo de Fondos"
    meses = list(pivot_final.columns)
    ncols = 1 + len(meses)
    _xl_setup_sheet(ws, "Flujo de Fondos · Criterio caja", periodo, ncols)
    ws.column_dimensions["A"].width = 22
    for i in range(len(meses)):
        ws.column_dimensions[_col(2 + i)].width = 16
    HR = 5
    _xl_table_header(ws, HR, ["Concepto"] + [str(m) for m in meses])
    ws.freeze_panes = f"B{HR + 1}"
    bold_rows = {"Saldo inicial", "Flujo neto", "Saldo final"}
    fill_rows = {"Saldo inicial", "Saldo final"}
    r = HR + 1
    for concepto in pivot_final.index:
        is_bold = concepto in bold_rows
        fill = _XL_NAVYSOFT if concepto in fill_rows else None
        top = concepto in ("Flujo neto", "Saldo final")
        _xl_text(ws, r, 1, str(concepto), bold=is_bold,
                 color=(_XL_NAVY if is_bold else _XL_DARK), fill=fill, top=top)
        for ci, m in enumerate(meses):
            _xl_money(ws, r, 2 + ci, float(pivot_final.loc[concepto, m]),
                      bold=is_bold, fill=fill, top=top)
        r += 1
    _xl_footer(ws, r + 1, ncols)
    out = io.BytesIO(); wb.save(out); return out.getvalue()

def _default_mes_idx(meses_ts):
    """Índice del mes en curso en la lista; si no está, el mes más reciente
    no futuro; si todos son futuros, el primero. Evita que los selectores
    arranquen en meses proyectados por cuotas."""
    hoy = date.today()
    mes_act = pd.Timestamp(hoy.year, hoy.month, 1)
    candidatos = [i for i, m in enumerate(meses_ts) if pd.Timestamp(m) <= mes_act]
    if not candidatos:
        return 0
    return max(candidatos, key=lambda i: pd.Timestamp(meses_ts[i]))

def rango_fechas_default(df, columna="fecha_devengo"):
    if df.empty:
        hoy = date.today()
        return hoy.replace(day=1), hoy
    hasta = df[columna].max().date()
    desde = (hasta - relativedelta(months=11)).replace(day=1)
    minimo = df[columna].min().date()
    if desde < minimo:
        desde = minimo.replace(day=1)
    return desde, hasta

# ============================================================================
# CÁLCULO DE ESTADO DE FLUJO (criterio caja)
# ============================================================================
def calcular_estado_flujo(user_id: str):
    """
    Calcula todo el estado de Flujo de Fondos del usuario en criterio caja.
    Se usa tanto en page_flujo (para la tabla detallada) como en page_dashboard
    (para los KPIs y gráficos).

    Devuelve None si el usuario no tiene movimientos. Si tiene, retorna un dict:
        {
            "df_caja":          DataFrame de cuotas expandidas
            "pivot_completo":   DataFrame pivot indexado por tipo, columnas = meses
            "flujo_neto":       Series indexada por mes con el flujo neto
            "meses_todos":      lista de Timestamps de meses con datos (ordenados)
            "ajustes_manuales": dict {date(yyyy,mm,01) -> ajuste_float}
            "saldo_auto":       dict {mes_ts -> saldo inicial sin ajuste}
            "ajuste":           dict {mes_ts -> ajuste aplicado al mes}
            "saldos_ini":       dict {mes_ts -> saldo inicial efectivo (auto + ajuste)}
            "saldos_fin":       dict {mes_ts -> saldo final (saldos_ini + flujo)}
        }
    """
    df = df_movimientos(user_id)
    if df.empty:
        return None
    df_caja = expandir_a_caja(df)
    if df_caja.empty:
        return None

    pivot_completo = df_caja.pivot_table(
        index="tipo", columns="mes_pago", values="monto_cuota",
        aggfunc="sum", fill_value=0,
    )
    orden = ["Ingreso", "Gasto Fijo", "Gasto Variable", "Ahorro"]
    pivot_completo = pivot_completo.reindex([t for t in orden if t in pivot_completo.index])

    # Devengado: total (monto_total) por tipo y mes de fecha_devengo.
    # Lo usan las vistas de "cuanto gane/gaste este mes" (KPIs del mes, ritmo y
    # torta) para coincidir con la pestana Movimientos. La caja se reserva para
    # Disponible real y para el Flujo de Fondos.
    df_dev = df.copy()
    df_dev["mes_dev"] = df_dev["fecha_devengo"].dt.to_period("M").dt.to_timestamp()
    pivot_dev = df_dev.pivot_table(
        index="tipo", columns="mes_dev", values="monto_total",
        aggfunc="sum", fill_value=0,
    )
    pivot_dev = pivot_dev.reindex([t for t in orden if t in pivot_dev.index])

    flujo_neto = (
        (pivot_completo.loc["Ingreso"] if "Ingreso" in pivot_completo.index else 0)
        - (pivot_completo.loc["Gasto Fijo"] if "Gasto Fijo" in pivot_completo.index else 0)
        - (pivot_completo.loc["Gasto Variable"] if "Gasto Variable" in pivot_completo.index else 0)
        - (pivot_completo.loc["Ahorro"] if "Ahorro" in pivot_completo.index else 0)
    )

    ajustes_manuales = get_saldos_iniciales_manuales(user_id)

    meses_todos = list(flujo_neto.index)
    saldo_auto, ajuste_map, saldos_ini, saldos_fin = {}, {}, {}, {}
    saldo_anterior = 0.0
    for mes_ts in meses_todos:
        mes_d = mes_ts.date().replace(day=1) if hasattr(mes_ts, "date") else mes_ts
        auto = saldo_anterior
        aj = float(ajustes_manuales.get(mes_d, 0.0))
        si = auto + aj
        sf = si + float(flujo_neto.loc[mes_ts])
        saldo_auto[mes_ts] = auto
        ajuste_map[mes_ts] = aj
        saldos_ini[mes_ts] = si
        saldos_fin[mes_ts] = sf
        saldo_anterior = sf

    return {
        "df_caja": df_caja,
        "df_dev": df_dev,
        "pivot_completo": pivot_completo,
        "pivot_dev": pivot_dev,
        "flujo_neto": flujo_neto,
        "meses_todos": meses_todos,
        "ajustes_manuales": ajustes_manuales,
        "saldo_auto": saldo_auto,
        "ajuste": ajuste_map,
        "saldos_ini": saldos_ini,
        "saldos_fin": saldos_fin,
    }

# ============================================================================
# UI HELPERS
# ============================================================================
def render_logo(ancho_px: int = 160):
    b64 = load_logo_b64(LOGO_PATH)
    if b64:
        st.markdown(
            f"<div style='text-align:center;'>"
            f"<img src='data:image/png;base64,{b64}' width='{ancho_px}' "
            f"alt='WL HNOS &amp; ASOC'/></div>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f"<div style='font-size:{ancho_px//3}px; text-align:center;'>💰</div>",
            unsafe_allow_html=True,
        )

def page_header(titulo: str, caption: str = ""):
    cap_html = f"<div class='page-caption'>{caption}</div>" if caption else ""
    st.markdown(
        f"<div class='page-header'>"
        f"<div class='page-title'>{titulo}</div>"
        f"{cap_html}"
        f"</div>",
        unsafe_allow_html=True,
    )

def metric_card(label: str, value: str, color: str = "", sublabel: str = ""):
    cls = f"metric-card metric-{color}" if color else "metric-card"
    sub = f"<div class='metric-sublabel'>{sublabel}</div>" if sublabel else ""
    return (
        f"<div class='{cls}'>"
        f"<div class='metric-label'>{label}</div>"
        f"<div class='metric-value'>{value}</div>"
        f"{sub}"
        f"</div>"
    )

def aplicar_tema_plotly(fig: go.Figure, height: int = 320):
    fig.update_layout(
        # es-AR: primer carácter = decimal, segundo = miles -> "1.000.000,00"
        separators=",.",
        font=dict(family="Open Sans, sans-serif", color=NAVY, size=12),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=10, r=10, t=10, b=10),
        height=height,
        legend=dict(
            orientation="h", yanchor="bottom", y=1.02, x=0,
            bgcolor="rgba(0,0,0,0)",
            font=dict(size=11),
        ),
        hoverlabel=dict(font_family="Open Sans, sans-serif", font_size=12),
    )
    fig.update_xaxes(
        showgrid=False,
        linecolor=BORDER,
        tickfont=dict(size=11, color=TEXT_MUTED),
    )
    fig.update_yaxes(
        gridcolor=BORDER,
        linecolor=BORDER,
        tickfont=dict(size=11, color=TEXT_MUTED),
        # Forzar número completo (sin 1k/1M/1G); junto con separators=",." -> "1.000.000"
        tickformat=",",
    )
    return fig

# ============================================================================
# PANTALLA: LOGIN
# ============================================================================
_GOOGLE_SVG = (
    "<svg width='18' height='18' viewBox='0 0 48 48'>"
    "<path fill='#FFC107' d='M43.6 20.5H42V20H24v8h11.3C33.7 32.7 29.2 36 24 36c-6.6 0-12-5.4-12-12s5.4-12 12-12c3.1 0 5.9 1.2 8 3l5.7-5.7C34.5 6.1 29.5 4 24 4 13 4 4 13 4 24s9 20 20 20 20-9 20-20c0-1.2-.1-2.3-.4-3.5z'/>"
    "<path fill='#FF3D00' d='M6.3 14.7l6.6 4.8C14.7 15.1 19 12 24 12c3.1 0 5.9 1.2 8 3l5.7-5.7C34.5 6.1 29.5 4 24 4 16.3 4 9.7 8.3 6.3 14.7z'/>"
    "<path fill='#4CAF50' d='M24 44c5.2 0 9.9-2 13.4-5.2l-6.2-5.2C29.2 35.1 26.7 36 24 36c-5.2 0-9.6-3.3-11.3-8l-6.5 5C9.5 39.6 16.2 44 24 44z'/>"
    "<path fill='#1976D2' d='M43.6 20.5H42V20H24v8h11.3c-.8 2.3-2.3 4.3-4.1 5.6l6.2 5.2C36.9 40.4 44 35 44 24c0-1.2-.1-2.3-.4-3.5z'/>"
    "</svg>"
)

def _pkce_par():
    """Genera (verifier, challenge) PKCE; el verifier persiste en la sesión."""
    if not st.session_state.get("pkce_verifier"):
        st.session_state.pkce_verifier = _secrets.token_urlsafe(64)
    v = st.session_state.pkce_verifier
    ch = base64.urlsafe_b64encode(hashlib.sha256(v.encode()).digest()).decode().rstrip("=")
    return v, ch

def page_login():
    _oe = st.session_state.pop("oauth_error", None)
    if _oe:
        st.error(f"No se pudo completar el ingreso con Google: {_oe}")
    col_a, col_b, col_c = st.columns([1, 2, 1])
    with col_b:
        render_logo(ancho_px=140)

    st.markdown(
        "<div style='text-align:center; margin-top:8px;'>"
        "<h1 style='margin-bottom:4px;'>Finanzas para protagonistas</h1>"
        f"<div style='color:{TEXT_MUTED}; font-size:0.9rem;'>"
        "Tus finanzas desde un solo lugar"
        "</div></div>",
        unsafe_allow_html=True,
    )

    st.write("")  # respiro entre header y tabs

    tab_in, tab_up = st.tabs(["Iniciar sesión", "Crear cuenta"])

    with tab_in:
        with st.form("login"):
            email = st.text_input("Email", placeholder="Escribe tu mail")
            password = st.text_input("Contraseña", type="password")
            ok = st.form_submit_button("Entrar", use_container_width=True)
            if ok:
                user, err = do_signin(email, password)
                if err:
                    if "Invalid login credentials" in err:
                        st.error("Email o contraseña incorrectos")
                    elif "Email not confirmed" in err:
                        st.error("Confirmá tu email antes de iniciar sesión.")
                    else:
                        st.error(f"Error: {err}")
                else:
                    st.session_state.user = user
                    st.rerun()

    with tab_up:
        with st.form("signup"):
            email = st.text_input("Email", key="su_email", placeholder="Escribe tu mail")
            password = st.text_input("Contraseña", type="password", key="su_pw",
                                     help="Mínimo 6 caracteres")
            password2 = st.text_input("Repetir contraseña", type="password", key="su_pw2")
            ok = st.form_submit_button("Crear cuenta", use_container_width=True)
            if ok:
                if not email or "@" not in email:
                    st.error("Email inválido")
                elif password != password2:
                    st.error("Las contraseñas no coinciden")
                elif len(password) < 6:
                    st.error("Mínimo 6 caracteres")
                else:
                    user, err = do_signup(email, password)
                    if err:
                        st.error(f"Error: {err}")
                    else:
                        st.success("Cuenta creada. Ya podés entrar desde la pestaña "
                                   "\"Iniciar sesión\" con tu mail y contraseña.")

    _verifier, _challenge = _pkce_par()
    # El verifier viaja en una cookie para que la pestaña que vuelve de Google
    # (que es una sesión nueva de Streamlit) pueda completar el intercambio.
    try:
        cookie_manager.set("wl_pkce", _verifier,
                           expires_at=datetime.now() + timedelta(minutes=15),
                           key="ck_pkce")
    except Exception:
        pass
    _auth_url = (f"{st.secrets['SUPABASE_URL'].strip().rstrip('/')}/auth/v1/authorize"
                 f"?provider=google&redirect_to={quote(APP_URL, safe='')}"
                 f"&code_challenge={_challenge}&code_challenge_method=s256")
    st.markdown(
        f"<a href='{_auth_url}' target='_blank' rel='noopener' class='google-btn'>"
        f"{_GOOGLE_SVG} Continuar con Google</a>",
        unsafe_allow_html=True,
    )
    st.write("")

    with st.expander("¿Olvidaste tu contraseña?"):
        with st.form("reset_pw"):
            email_r = st.text_input("Email", key="rp_email", placeholder="Escribe tu mail")
            ok_r = st.form_submit_button("Enviarme el link de recuperación",
                                         use_container_width=True)
            if ok_r:
                if not email_r or "@" not in email_r:
                    st.error("Escribí tu mail")
                else:
                    try:
                        sb.auth.reset_password_for_email(email_r)
                    except Exception:
                        pass
                    # Mismo mensaje exista o no la cuenta (no revelar emails registrados)
                    st.success("Si esa cuenta existe, te enviamos un mail con el link "
                               "para crear una nueva contraseña. Revisá también spam.")

    st.markdown(
        "<div class='login-footer'>WL HNOS &amp; ASOC · Catamarca · "
        "<a href='https://wlhnos.vercel.app/' target='_blank' "
        "style='color: var(--cyan); text-decoration: none;'>wlhnos.vercel.app</a></div>",
        unsafe_allow_html=True,
    )

# ============================================================================
# PANTALLA: NUEVA CONTRASEÑA (llega desde el link del mail de recuperación)
# ============================================================================
def page_nueva_password():
    col_a, col_b, col_c = st.columns([1, 2, 1])
    with col_b:
        render_logo(ancho_px=150)
        st.markdown(
            "<div style='text-align:center; margin-top:8px;'>"
            "<h1 style='margin-bottom:4px;'>Crear nueva contraseña</h1></div>",
            unsafe_allow_html=True,
        )
        with st.form("nueva_pw"):
            p1 = st.text_input("Nueva contraseña", type="password",
                               help="Mínimo 6 caracteres")
            p2 = st.text_input("Repetir contraseña", type="password")
            ok = st.form_submit_button("Guardar y entrar", use_container_width=True)
        if ok:
            if len(p1) < 6:
                st.error("Mínimo 6 caracteres")
            elif p1 != p2:
                st.error("Las contraseñas no coinciden")
            else:
                try:
                    r = sb.auth.verify_otp({
                        "token_hash": st.session_state.recovery_token_hash,
                        "type": "recovery",
                    })
                    if r and r.session:
                        sb.auth.update_user({"password": p1})
                        s = sb.auth.get_session()
                        tok = s or r.session
                        st.session_state.sb_tokens = {
                            "access_token": tok.access_token,
                            "refresh_token": tok.refresh_token,
                        }
                        st.session_state.user = r.user
                        st.session_state.recovery_token_hash = None
                        st.rerun()
                    else:
                        st.error("El link no es válido. Pedí uno nuevo desde el login.")
                except Exception:
                    st.session_state.recovery_token_hash = None
                    st.error("El link venció o ya fue usado. Volvé al login y pedí "
                             "uno nuevo desde la opción de contraseña olvidada.")
        if st.button("Volver al inicio de sesión", use_container_width=True):
            st.session_state.recovery_token_hash = None
            st.rerun()

# ============================================================================
# PANTALLA: CARGAR MOVIMIENTO
# ============================================================================
def _opciones_categorias(user_id, grupo):
    """
    {label: (nombre_categoria, tipo_db)} para el grupo elegido.
    grupo "Gasto" une Gasto Fijo + Gasto Variable; la categoría elegida
    determina el tipo que se guarda (el usuario ya no clasifica a mano).
    """
    if grupo == "Gasto":
        out = {}
        for tipo_db, sufijo in [("Gasto Fijo", "fijo"), ("Gasto Variable", "variable")]:
            for c in get_categorias(user_id, tipo_db):
                out[f"{c['nombre']} · {sufijo}"] = (c["nombre"], tipo_db)
        return dict(sorted(out.items()))
    return {c["nombre"]: (c["nombre"], grupo) for c in get_categorias(user_id, grupo)}

def page_cargar(user):
    page_header("Nuevo movimiento",
                "Cargá ingresos, gastos o ahorro con criterio devengado y caja.")

    grupo = st.selectbox(
        "Tipo",
        ["Ingreso", "Gasto", "Ahorro"],
        index=1,
        key="tipo_select",
    )
    opciones_cat = _opciones_categorias(user.id, grupo)

    # Fuera del form para que muestre/oculte los campos de cuotas al instante.
    en_cuotas = st.toggle(
        "Pago en cuotas o en otra fecha",
        value=False,
        key="cargar_en_cuotas",
        help="Activalo solo si el pago no es el mismo día del movimiento "
             "(cuotas, tarjeta, pago diferido).",
    )

    with st.form("nuevo_mov", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            fecha_devengo = st.date_input("Fecha del movimiento", value=date.today(),
                                          format="DD/MM/YYYY")
        with col2:
            if not opciones_cat:
                st.warning(f"Sin categorías de tipo '{grupo}'. Creá una en Configuración.")
                cat_lbl = None
            else:
                cat_lbl = st.selectbox("Categoría", list(opciones_cat.keys()))

        concepto = st.text_input("Concepto / Nota", placeholder="Ej: Período 05/26")
        monto = st.number_input("Monto total", min_value=0.0, step=1000.0,
                                format="%.2f", value=None, placeholder="0,00")

        if en_cuotas:
            col3, col4 = st.columns(2)
            with col3:
                cuotas = st.number_input("Cuotas", min_value=1, max_value=36, value=1, step=1)
            with col4:
                inicio_pago = st.date_input("Inicio pago (primera cuota)", value=date.today(),
                                            format="DD/MM/YYYY")
            st.caption(
                "El movimiento completo impacta en el Estado de Resultados en su fecha "
                "(devengado); las cuotas se reparten en el Flujo de Fondos desde el inicio "
                "de pago (caja)."
            )
        else:
            cuotas = 1
            inicio_pago = None  # se usa la fecha del movimiento

        ok = st.form_submit_button("Guardar movimiento", use_container_width=True)
        if ok:
            if not cat_lbl:
                st.error("Elegí una categoría")
            elif not monto or monto <= 0:
                st.error("El monto debe ser mayor a cero")
            else:
                try:
                    categoria, tipo = opciones_cat[cat_lbl]
                    inicio_efectivo = inicio_pago if en_cuotas else fecha_devengo
                    insert_movimiento(user.id, fecha_devengo, tipo, categoria, concepto,
                                      monto, cuotas, inicio_efectivo)
                    st.success(f"{tipo} de {fmt_money(monto)} guardado")
                    st.cache_data.clear()
                except Exception as e:
                    st.error(f"Error al guardar: {e}")

    st.divider()
    with st.expander("Copiar fijos e ingresos del mes anterior"):
        if st.session_state.pop("flash_copia", None):
            st.success("Movimientos copiados")
        hoy = date.today()
        per_dest = pd.Period(date(hoy.year, hoy.month, 1), "M")
        per_orig = per_dest - 1
        df_all = df_movimientos(user.id)
        base = pd.DataFrame()
        if not df_all.empty:
            base = df_all[
                df_all["tipo"].isin(["Gasto Fijo", "Ingreso"])
                & (df_all["fecha_devengo"].dt.to_period("M") == per_orig)
                & (df_all["cuotas"] == 1)
            ].copy()
        if base.empty:
            st.caption(
                f"No hay gastos fijos ni ingresos (sin cuotas) cargados en "
                f"{per_orig.strftime('%m/%Y')} para copiar."
            )
        else:
            ya = df_all[df_all["fecha_devengo"].dt.to_period("M") == per_dest]
            claves_ya = set(zip(ya["tipo"], ya["categoria"], ya["concepto"].fillna("")))
            base = base.sort_values(["tipo", "categoria"]).reset_index(drop=True)
            base["_clave"] = list(zip(base["tipo"], base["categoria"],
                                      base["concepto"].fillna("")))
            base["Cargar"] = ~base["_clave"].isin(claves_ya)

            ed = pd.DataFrame({
                "Cargar": base["Cargar"],
                "Tipo": base["tipo"],
                "Categoría": base["categoria"],
                "Concepto": base["concepto"].fillna(""),
                "Monto": base["monto_total"],
                "Fecha nueva": (base["fecha_devengo"] + pd.DateOffset(months=1))
                                .dt.strftime("%d/%m/%Y"),
            })
            st.caption(
                f"Gastos fijos e ingresos de **{per_orig.strftime('%m/%Y')}** → "
                f"**{per_dest.strftime('%m/%Y')}**. Editá montos (aumentos) y conceptos, "
                f"y destildá lo que no corresponda. Los que parecen ya cargados este mes "
                f"vienen destildados."
            )
            ed_out = st.data_editor(
                ed, hide_index=True, use_container_width=True,
                disabled=["Tipo", "Categoría", "Fecha nueva"],
                key="copiar_fijos_editor",
            )
            n_sel = int(ed_out["Cargar"].sum())
            if st.button(f"Cargar {n_sel} movimiento(s) en {per_dest.strftime('%m/%Y')}",
                         disabled=n_sel == 0, key="btn_copiar_fijos"):
                fechas_nuevas = (base["fecha_devengo"] + pd.DateOffset(months=1)).dt.date.tolist()
                errores = 0
                for (_, row), fnueva in zip(ed_out.iterrows(), fechas_nuevas):
                    if not row["Cargar"] or not row["Monto"] or float(row["Monto"]) <= 0:
                        continue
                    try:
                        insert_movimiento(user.id, fnueva, row["Tipo"], row["Categoría"],
                                          row["Concepto"], float(row["Monto"]), 1, fnueva)
                    except Exception:
                        errores += 1
                st.cache_data.clear()
                if errores:
                    st.warning(f"Hubo {errores} error(es) al copiar; revisá la lista.")
                else:
                    st.session_state["flash_copia"] = True
                st.rerun()

    st.divider()
    st.markdown("##### Últimos movimientos")
    movs = get_movimientos(user.id, limit=10)
    if not movs:
        st.caption("Todavía no cargaste ningún movimiento.")
        return

    cls_map = {
        "Ingreso": "mov-ingreso",
        "Gasto Fijo": "mov-fijo",
        "Gasto Variable": "mov-variable",
        "Ahorro": "mov-ahorro",
    }
    for m in movs:
        cls = cls_map.get(m["tipo"], "")
        sign = "+ " if m["tipo"] == "Ingreso" else ""
        cuotas_lbl = f" · {m['cuotas']} cuotas" if m["cuotas"] > 1 else ""
        fecha_fmt = pd.to_datetime(m["fecha_devengo"]).strftime("%d/%m/%Y")
        concepto_lbl = m["concepto"] or m["tipo"]
        st.markdown(
            f"<div class='mov-row {cls}'>"
            f"  <div class='mov-info'>"
            f"    <div class='mov-cat'>{m['categoria']}</div>"
            f"    <div class='mov-desc'>{concepto_lbl} · {fecha_fmt}{cuotas_lbl}</div>"
            f"  </div>"
            f"  <div class='mov-monto'>{sign}{fmt_money(m['monto_total'])}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

# ============================================================================
# PANTALLA: VER / EDITAR / BORRAR MOVIMIENTOS
# ============================================================================
def page_ver(user):
    page_header("Movimientos",
                "Tus movimientos cargados. Podés editar o eliminar abajo.")

    df = df_movimientos(user.id)
    if df.empty:
        st.info("No tenés movimientos cargados todavía.")
        return

    # --- Filtros ---
    meses_disp = sorted(df["fecha_devengo"].dt.strftime("%m/%Y").unique(),
                        key=lambda s: (s[3:], s[:2]), reverse=True)
    cats_disp = sorted(df["categoria"].unique())
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        mes_f = st.selectbox("Mes", ["Todos"] + meses_disp, key="ver_mes_f")
    with col_f2:
        tipo_f = st.selectbox("Tipo", ["Todos", "Ingreso", "Gasto Fijo",
                                       "Gasto Variable", "Ahorro"], key="ver_tipo_f")
    with col_f3:
        cat_f = st.selectbox("Categoría", ["Todas"] + cats_disp, key="ver_cat_f")
    if mes_f != "Todos":
        df = df[df["fecha_devengo"].dt.strftime("%m/%Y") == mes_f]
    if tipo_f != "Todos":
        df = df[df["tipo"] == tipo_f]
    if cat_f != "Todas":
        df = df[df["categoria"] == cat_f]
    if df.empty:
        st.info("No hay movimientos con esos filtros.")
        return

    total_ing = df.loc[df["tipo"] == "Ingreso", "monto_total"].sum()
    total_gas = df.loc[df["tipo"].isin(["Gasto Fijo", "Gasto Variable"]), "monto_total"].sum()
    total_aho = df.loc[df["tipo"] == "Ahorro", "monto_total"].sum()
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(metric_card("Ingresos", fmt_money(total_ing)),
                    unsafe_allow_html=True)
    with col2:
        st.markdown(metric_card("Gastos", fmt_money(total_gas), "orange"),
                    unsafe_allow_html=True)
    with col3:
        st.markdown(metric_card("Ahorro", fmt_money(total_aho), "cyan"),
                    unsafe_allow_html=True)

    st.write("")

    df_view = df.copy()
    df_view["concepto"] = df_view["concepto"].fillna("")
    df_view["Fecha"] = df_view["fecha_devengo"].dt.strftime("%d/%m/%Y")
    df_view["Inicio pago"] = df_view["inicio_pago"].dt.strftime("%d/%m/%Y")
    df_view["Monto"] = df_view["monto_total"].apply(fmt_money)
    df_view = df_view[["Fecha", "tipo", "categoria", "concepto",
                       "Monto", "cuotas", "Inicio pago"]]
    df_view.columns = ["Fecha", "Tipo", "Categoría", "Concepto",
                       "Monto", "Cuotas", "Inicio pago"]
    st.dataframe(df_view, hide_index=True, use_container_width=True)

    # --- Exportar a Excel (respeta los filtros activos) ---  [PREMIUM]
    if es_premium(user):
        _partes = [f"Mes: {mes_f}" if mes_f != "Todos" else "Todos los meses"]
        if tipo_f != "Todos":
            _partes.append(f"Tipo: {tipo_f}")
        if cat_f != "Todas":
            _partes.append(f"Categoría: {cat_f}")
        _periodo_lbl = "   ·   ".join(_partes)
        try:
            _xlsx_mov = build_xlsx_movimientos(df, _periodo_lbl)
            st.download_button(
                "Descargar Excel (Resumen + Detalle)",
                data=_xlsx_mov,
                file_name="movimientos_por_categoria.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.caption(f"No se pudo generar el Excel: {e}")
    else:
        boton_premium("Descargar Excel (Resumen + Detalle)", key="lock_xlsx_mov")

    st.divider()
    st.markdown("##### Editar / Eliminar")

    movs = df.to_dict("records")
    def label_mov(m):
        f = pd.to_datetime(m["fecha_devengo"]).strftime("%d/%m/%Y")
        return f"{f} · {m['tipo']} · {m['categoria']} · {fmt_money(m['monto_total'])} · {m['concepto'] or ''}"

    opciones = {label_mov(m): m["id"] for m in movs}
    label_sel = st.selectbox(
        "Elegí el movimiento",
        options=list(opciones.keys()),
        index=None,
        placeholder="Elegir movimiento…",
    )

    if not label_sel:
        return

    mov_id = opciones[label_sel]
    mov = next(m for m in movs if m["id"] == mov_id)

    st.caption(f"Editando: {label_sel}")

    # El selector de Tipo va FUERA del form: los widgets dentro de un st.form
    # no disparan rerun. Tipo simplificado a Ingreso/Gasto/Ahorro: la categoría
    # determina si un gasto es Fijo o Variable.
    grupos = ["Ingreso", "Gasto", "Ahorro"]
    grupo_mov = "Gasto" if mov["tipo"] in ("Gasto Fijo", "Gasto Variable") else mov["tipo"]
    grupo_e = st.selectbox("Tipo", grupos, index=grupos.index(grupo_mov),
                           key=f"e_tipo_{mov_id}")

    opciones_cat = _opciones_categorias(user.id, grupo_e)
    # Ofrecer la categoría original aunque esté inactiva (solo si el grupo coincide).
    lbl_orig = None
    if grupo_e == grupo_mov:
        if grupo_mov == "Gasto":
            sufijo = "fijo" if mov["tipo"] == "Gasto Fijo" else "variable"
            lbl_orig = f"{mov['categoria']} · {sufijo}"
        else:
            lbl_orig = mov["categoria"]
        if lbl_orig not in opciones_cat:
            opciones_cat = {lbl_orig: (mov["categoria"], mov["tipo"]), **opciones_cat}

    if not opciones_cat:
        st.warning(f"Sin categorías activas de tipo '{grupo_e}'. Creá una en Configuración.")

    with st.form(f"edit_{mov_id}"):
        col1, col2 = st.columns(2)
        with col1:
            fecha_e = st.date_input("Fecha del movimiento",
                                    value=pd.to_datetime(mov["fecha_devengo"]).date(),
                                    format="DD/MM/YYYY", key=f"e_fdev_{mov_id}")
        with col2:
            if opciones_cat:
                labels = list(opciones_cat.keys())
                cat_idx = labels.index(lbl_orig) if lbl_orig in labels else 0
                cat_lbl_e = st.selectbox("Categoría", labels, index=cat_idx,
                                         key=f"e_cat_{mov_id}")
            else:
                cat_lbl_e = None

        concepto_e = st.text_input("Concepto / Nota",
                                   value=mov["concepto"] or "", key=f"e_con_{mov_id}")
        monto_e = st.number_input("Monto total",
                                  min_value=0.0, step=1000.0, format="%.2f",
                                  value=float(mov["monto_total"]), key=f"e_mon_{mov_id}")

        col3, col4 = st.columns(2)
        with col3:
            cuotas_e = st.number_input("Cuotas", min_value=1, max_value=36,
                                       value=int(mov["cuotas"]), step=1, key=f"e_cuo_{mov_id}")
        with col4:
            inicio_e = st.date_input("Inicio pago",
                                     value=pd.to_datetime(mov["inicio_pago"]).date(),
                                     format="DD/MM/YYYY", key=f"e_inip_{mov_id}")

        col_g, col_d = st.columns(2)
        with col_g:
            guardar = st.form_submit_button("Guardar cambios", use_container_width=True)
        with col_d:
            eliminar = st.form_submit_button("Eliminar", use_container_width=True)

        if guardar:
            if not cat_lbl_e:
                st.error("Elegí una categoría")
            elif monto_e <= 0:
                st.error("El monto debe ser mayor a cero")
            else:
                try:
                    cat_e, tipo_db_e = opciones_cat[cat_lbl_e]
                    update_movimiento(user.id, mov_id, fecha_e, tipo_db_e, cat_e,
                                      concepto_e, monto_e, cuotas_e, inicio_e)
                    st.success("Movimiento actualizado")
                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al actualizar: {e}")

        if eliminar:
            st.session_state[f"confirmar_del_{mov_id}"] = True
            st.rerun()

    if st.session_state.get(f"confirmar_del_{mov_id}"):
        st.warning(f"¿Confirmás eliminar este movimiento? **{label_sel}**")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Sí, eliminar", use_container_width=True, key=f"si_del_{mov_id}"):
                try:
                    delete_movimiento(user.id, mov_id)
                    st.session_state.pop(f"confirmar_del_{mov_id}", None)
                    st.success("Movimiento eliminado")
                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al eliminar: {e}")
        with c2:
            if st.button("Cancelar", use_container_width=True, key=f"no_del_{mov_id}"):
                st.session_state.pop(f"confirmar_del_{mov_id}", None)
                st.rerun()

# ============================================================================
# PANTALLA: ESTADO DE RESULTADOS
# ============================================================================
def _reporte_resultados(user):
    st.caption("Criterio devengado · Agrupado por el mes de la fecha del movimiento.")

    df = df_movimientos(user.id)
    if df.empty:
        st.info("No tenés movimientos cargados todavía.")
        return

    desde_def, hasta_def = rango_fechas_default(df, "fecha_devengo")
    fechas = st.date_input(
        "Rango de fechas",
        value=(desde_def, hasta_def),
        format="DD/MM/YYYY",
        key="rango_er",
    )
    if isinstance(fechas, tuple) and len(fechas) == 2:
        desde, hasta = fechas
    else:
        desde, hasta = desde_def, hasta_def

    df = df[
        (df["fecha_devengo"].dt.date >= desde) &
        (df["fecha_devengo"].dt.date <= hasta)
    ].copy()

    if df.empty:
        st.info("No hay movimientos en el rango seleccionado.")
        return

    df["mes"] = df["fecha_devengo"].dt.to_period("M").dt.to_timestamp()

    pivot = df.pivot_table(
        index="tipo", columns="mes", values="monto_total",
        aggfunc="sum", fill_value=0,
    )
    orden = ["Ingreso", "Gasto Fijo", "Gasto Variable", "Ahorro"]
    pivot = pivot.reindex([t for t in orden if t in pivot.index])

    def _fila(t):
        return pivot.loc[t] if t in pivot.index else 0
    resultado = _fila("Ingreso") - _fila("Gasto Fijo") - _fila("Gasto Variable")
    pivot.loc["Resultado del período"] = resultado
    pivot.loc["Resultado después de ahorro"] = resultado - _fila("Ahorro")
    orden_final = [r for r in ["Ingreso", "Gasto Fijo", "Gasto Variable",
                               "Resultado del período", "Ahorro",
                               "Resultado después de ahorro"] if r in pivot.index]
    pivot = pivot.reindex(orden_final)

    pivot.columns = [c.strftime("%m/%Y") for c in pivot.columns]
    pivot_fmt = pivot.copy().map(fmt_money)
    pivot_fmt.index.name = "Concepto"
    st.dataframe(pivot_fmt, use_container_width=True)

    st.caption(
        "Los gráficos de evolución mensual y de gastos por categoría viven en **Inicio**."
    )

    st.markdown("##### Detalle por categoría")
    meses_disp = sorted(df["mes"].unique(), reverse=True)
    mes_sel = st.selectbox(
        "Seleccionar mes",
        options=meses_disp,
        index=_default_mes_idx(meses_disp),
        format_func=lambda d: pd.Timestamp(d).strftime("%m/%Y"),
        key="mes_detalle_er",
    )
    df_mes = df[df["mes"] == mes_sel]
    detalle = df_mes.groupby(["tipo", "categoria"], as_index=False)["monto_total"].sum()
    detalle.columns = ["Tipo", "Categoría", "Monto"]
    detalle["Monto"] = detalle["Monto"].apply(fmt_money)
    st.dataframe(detalle, hide_index=True, use_container_width=True)

# ============================================================================
# PANTALLA: FLUJO DE FONDOS
# ============================================================================
def _reporte_flujo(user):
    st.caption("Criterio caja · Cuotas distribuidas en el mes que corresponde.")

    df = df_movimientos(user.id)
    if df.empty:
        st.info("No tenés movimientos cargados todavía.")
        return

    # ------------------------------------------------------------------------
    # CÁLCULO COMPARTIDO con page_dashboard (sobre TODOS los meses)
    # ------------------------------------------------------------------------
    estado = calcular_estado_flujo(user.id)
    if estado is None:
        st.info("No hay cuotas que mostrar.")
        return
    df_caja = estado["df_caja"]
    pivot_completo = estado["pivot_completo"]
    flujo_neto_completo = estado["flujo_neto"]
    meses_todos = estado["meses_todos"]
    ajustes_manuales = estado["ajustes_manuales"]
    saldo_auto = estado["saldo_auto"]
    ajuste = estado["ajuste"]
    saldos_ini = estado["saldos_ini"]
    saldos_fin = estado["saldos_fin"]

    # --- Rango de fechas por default (últimos 12 meses) ---
    hasta_def = df_caja["mes_pago"].max().date()
    desde_def = (hasta_def - relativedelta(months=11)).replace(day=1)
    minimo = df_caja["mes_pago"].min().date()
    if desde_def < minimo:
        desde_def = minimo.replace(day=1)

    fechas = st.date_input(
        "Rango de fechas (mes de pago)",
        value=(desde_def, hasta_def),
        format="DD/MM/YYYY",
        key="rango_ff",
    )
    if isinstance(fechas, tuple) and len(fechas) == 2:
        desde, hasta = fechas
    else:
        desde, hasta = desde_def, hasta_def

    # ------------------------------------------------------------------------
    # Filtrar al rango pedido para mostrar
    # ------------------------------------------------------------------------
    meses_rango = [m for m in meses_todos
                   if desde <= (m.date() if hasattr(m, "date") else m) <= hasta]
    if not meses_rango:
        st.info("No hay cuotas en el rango seleccionado.")
        return

    pivot_rango = pivot_completo.loc[:, meses_rango].copy()

    # ------------------------------------------------------------------------
    # Armar tabla final
    # Filas: Saldo inicial → [Ajuste si hay alguno en rango] → tipos → Flujo neto → Saldo final
    # ------------------------------------------------------------------------
    hay_ajustes_en_rango = any(ajuste[m] != 0.0 for m in meses_rango)
    filas_ordenadas = ["Saldo inicial"]
    if hay_ajustes_en_rango:
        filas_ordenadas.append("Ajuste")
    filas_ordenadas += list(pivot_rango.index) + ["Flujo neto", "Saldo final"]

    pivot_final = pd.DataFrame(index=filas_ordenadas, columns=pivot_rango.columns, dtype=float)

    for m in meses_rango:
        pivot_final.loc["Saldo inicial", m] = saldos_ini[m]
        if hay_ajustes_en_rango:
            pivot_final.loc["Ajuste", m] = ajuste[m]
        for tipo in pivot_rango.index:
            pivot_final.loc[tipo, m] = pivot_rango.loc[tipo, m]
        pivot_final.loc["Flujo neto", m] = float(flujo_neto_completo.loc[m])
        pivot_final.loc["Saldo final", m] = saldos_fin[m]

    pivot_final.columns = [c.strftime("%m/%Y") for c in pivot_final.columns]
    pivot_fmt = pivot_final.copy().map(fmt_money)
    pivot_fmt.index.name = "Concepto"
    st.dataframe(pivot_fmt, use_container_width=True)

    # Aviso si el primer mes con datos no tiene ajuste cargado (arranca en $0)
    primer_mes_global = meses_todos[0]
    primer_mes_global_d = (primer_mes_global.date().replace(day=1)
                           if hasattr(primer_mes_global, "date") else primer_mes_global)
    if primer_mes_global_d not in ajustes_manuales:
        primer_lbl = primer_mes_global.strftime("%m/%Y")
        st.caption(
            f"El primer mes con movimientos ({primer_lbl}) arranca con saldo inicial $0,00. "
            f"Si tenías dinero antes de empezar a usar la app, conciliá el saldo de ese mes "
            f"abajo: decile a la app cuánto tenías realmente y listo."
        )

    # ------------------------------------------------------------------------
    # AJUSTAR saldo inicial de un mes (widgets sueltos para preview en vivo)
    # ------------------------------------------------------------------------
    st.divider()
    st.markdown("##### Conciliar saldo")
    st.caption(
        "Elegí un mes y decile a la app cuál es tu saldo REAL a fin de ese mes "
        "(lo que ves en tu billetera o banco). La diferencia se registra como "
        "ajuste y se arrastra a los meses siguientes."
    )

    meses_opts_lbl = [m.strftime("%m/%Y") for m in meses_rango]
    meses_opts_ts = list(meses_rango)

    col_mes, col_saldo = st.columns([1, 2])
    with col_mes:
        mes_idx = st.selectbox(
            "Mes",
            options=list(range(len(meses_opts_lbl))),
            format_func=lambda i: meses_opts_lbl[i],
            index=_default_mes_idx(meses_opts_ts),
            key="conc_mes_idx",
        )
    mes_ts_sel = meses_opts_ts[mes_idx]
    mes_lbl_sel = meses_opts_lbl[mes_idx]
    saldo_fin_actual = float(saldos_fin[mes_ts_sel])
    ajuste_actual = float(ajuste[mes_ts_sel])

    with col_saldo:
        # Key dinámica por mes: al cambiar el selector, el input se resetea
        # al saldo calculado de ese mes.
        saldo_real = st.number_input(
            f"Tu saldo real a fin de {mes_lbl_sel}",
            value=saldo_fin_actual,
            step=10000.0,
            format="%.2f",
            key=f"conc_input_{mes_ts_sel.isoformat()}",
        )

    delta = float(saldo_real or 0.0) - saldo_fin_actual
    if abs(delta) < 0.005:
        st.caption(
            f"✅ Coincide con el saldo calculado por la app "
            f"({fmt_money(saldo_fin_actual)}). No hay nada que ajustar."
        )
    else:
        st.markdown(
            f"<div style='background: var(--bg-soft); border: 1px solid var(--border); "
            f"border-radius: 8px; padding: 12px 16px; margin: 8px 0;'>"
            f"<div style='color: var(--text-muted); font-size: 0.85rem;'>Vista previa para {mes_lbl_sel}:</div>"
            f"<div style='font-family: Poppins, sans-serif; color: var(--navy); margin-top: 4px;'>"
            f"Según la app: {fmt_money(saldo_fin_actual)} · Tu saldo real: {fmt_money(float(saldo_real or 0.0))} · "
            f"Ajuste a registrar: <strong>{fmt_money(delta)}</strong>"
            f"</div></div>",
            unsafe_allow_html=True,
        )

    if st.button("Guardar conciliación", key="btn_conciliar", disabled=abs(delta) < 0.005):
        try:
            nuevo_ajuste = ajuste_actual + delta
            if abs(nuevo_ajuste) < 0.005:
                delete_saldo_inicial(user.id, mes_ts_sel.date())
                st.success(f"Conciliado: {mes_lbl_sel} vuelve al cálculo automático.")
            else:
                upsert_saldo_inicial(user.id, mes_ts_sel.date(), nuevo_ajuste)
                st.success(f"Conciliado: saldo de {mes_lbl_sel} = {fmt_money(float(saldo_real))}.")
            st.cache_data.clear()
            st.rerun()
        except Exception as e:
            st.error(f"Error al guardar: {e}")

    # ------------------------------------------------------------------------
    # Detalle de cuotas por mes (igual que antes)
    # ------------------------------------------------------------------------
    st.divider()
    st.markdown("##### Detalle de cuotas por mes")
    df_caja_rango = df_caja[
        (df_caja["mes_pago"].dt.date >= desde) &
        (df_caja["mes_pago"].dt.date <= hasta)
    ].copy()
    meses_det = sorted(df_caja_rango["mes_pago"].unique(), reverse=True)
    mes_sel = st.selectbox(
        "Seleccionar mes",
        options=meses_det,
        index=_default_mes_idx(meses_det),
        format_func=lambda d: pd.Timestamp(d).strftime("%m/%Y"),
        key="flujo_mes_detalle",
    )
    df_mes = df_caja_rango[df_caja_rango["mes_pago"] == mes_sel].copy()
    df_mes["concepto"] = df_mes["concepto"].fillna("")
    df_mes["Cuota"] = df_mes.apply(lambda r: f"{r['n_cuota']}/{r['total_cuotas']}", axis=1)
    df_mes["Monto"] = df_mes["monto_cuota"].apply(fmt_money)
    df_mes = df_mes[["tipo", "categoria", "concepto", "Cuota", "Monto"]]
    df_mes.columns = ["Tipo", "Categoría", "Concepto", "Cuota", "Monto"]
    st.dataframe(df_mes, hide_index=True, use_container_width=True)

    # [PREMIUM] Descarga del Flujo de Fondos en Excel
    if es_premium(user):
        _periodo_ff = f"{desde.strftime('%m/%Y')} a {hasta.strftime('%m/%Y')}"
        try:
            _xlsx_ff = build_xlsx_flujo(pivot_final, _periodo_ff)
            st.download_button(
                "Descargar Flujo de Fondos (Excel)",
                data=_xlsx_ff,
                file_name="flujo_de_fondos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.caption(f"No se pudo generar el Excel: {e}")
    else:
        boton_premium("Descargar Flujo de Fondos (Excel)", key="lock_xlsx_ff")

# ============================================================================
# PANTALLA: REPORTES (Estado de Resultados + Flujo de Fondos)
# ============================================================================
def page_reportes(user):
    page_header("Reportes", "Dos miradas sobre los mismos datos: devengado y caja.")
    criterio = st.radio(
        "Criterio",
        ["Estado de Resultados", "Flujo de Fondos"],
        horizontal=True,
        key="rep_criterio",
        label_visibility="collapsed",
    )
    with st.expander("¿Cuál es la diferencia entre los dos?"):
        st.markdown(
            "**Estado de Resultados (criterio devengado)**\n\n"
            "Mira el mes en el que ocurrió el hecho económico, sin importar cuándo "
            "lo pagaste. Si comprás algo en 12 cuotas en mayo, el gasto entero "
            "aparece en mayo. Responde: *¿qué tan bien me fue este mes?*\n\n"
            "**Flujo de Fondos (criterio caja)**\n\n"
            "Mira el mes en el que efectivamente entra o sale el dinero. Si comprás "
            "algo en 12 cuotas en mayo, cada cuota aparece en su mes. Responde: "
            "*¿cuánto dinero tengo / voy a tener cada mes?*"
        )
    st.write("")
    if "Resultados" in criterio:
        _reporte_resultados(user)
    else:
        _reporte_flujo(user)

# ============================================================================
# PANTALLA: CONFIGURACIÓN DE CATEGORÍAS
# ============================================================================
def _footer_sesion(user):
    st.divider()
    col_u, col_b = st.columns([3, 1])
    with col_u:
        st.caption(f"Sesión: {user.email}")
    with col_b:
        if st.button("Cerrar sesión", key="signout_footer", use_container_width=True):
            do_signout()

def page_configuracion(user):
    page_header("Configuración",
                "Tu nombre y las categorías que usás al cargar movimientos.")

    # ---- Tu nombre (cómo te saluda la app) ----
    st.markdown("##### Tu nombre")
    _meta = getattr(user, "user_metadata", None) or {}
    _nombre_actual = (_meta.get("display_name") or "").strip()
    _cn, _cb = st.columns([3, 1])
    with _cn:
        _nuevo_nombre = st.text_input(
            "Nombre", value=_nombre_actual, max_chars=40,
            placeholder="Coloca tu nombre",
            label_visibility="collapsed", key="cfg_display_name",
        )
    with _cb:
        _guardar_nombre = st.button("Guardar", use_container_width=True, key="cfg_save_name")
    if _guardar_nombre:
        try:
            _resp = sb.auth.update_user({"data": {"display_name": (_nuevo_nombre or "").strip()}})
            if _resp and getattr(_resp, "user", None):
                st.session_state.user = _resp.user
            st.success("Nombre actualizado.")
            st.rerun()
        except Exception as e:
            st.error(f"No se pudo guardar: {e}")

    st.divider()
    st.markdown("##### Categorías")
    st.caption("Si desactivás una categoría deja de aparecer en el selector, "
               "pero los movimientos viejos se conservan.")

    # [PLANES] Plan del usuario y cuántas categorías propias lleva creadas.
    premium = es_premium(user)
    propias = contar_categorias_propias(user.id)
    if not premium:
        st.caption(
            f"🔒 Plan gratis: podés crear hasta {LIMITE_CATEGORIAS_PROPIAS_GRATIS} "
            f"categorías propias en total ({propias} usadas). Renombrar, desactivar "
            f"y categorías ilimitadas son parte de **Premium**."
        )

    tipo_sel = st.selectbox(
        "Tipo",
        ["Ingreso", "Gasto Fijo", "Gasto Variable", "Ahorro"],
        key="cfg_tipo",
    )

    _puede_agregar = premium or propias < LIMITE_CATEGORIAS_PROPIAS_GRATIS
    if _puede_agregar:
        with st.form(f"nueva_cat_{tipo_sel}", clear_on_submit=True):
            st.markdown("##### Agregar categoría")
            col1, col2 = st.columns([3, 1])
            with col1:
                nuevo_nombre = st.text_input(
                    "Nombre",
                    placeholder=f"Ej: Internet (categoría {tipo_sel.lower()})",
                    label_visibility="collapsed",
                )
            with col2:
                ok_nueva = st.form_submit_button("Agregar", use_container_width=True)
            if ok_nueva:
                n = (nuevo_nombre or "").strip()
                if not n:
                    st.error("Ingresá un nombre")
                elif categoria_existe(user.id, n, tipo_sel):
                    st.error(f"Ya existe la categoría '{n}' para tipo {tipo_sel}")
                elif (not premium
                      and contar_categorias_propias(user.id) >= LIMITE_CATEGORIAS_PROPIAS_GRATIS):
                    st.warning(
                        f"Llegaste al límite de {LIMITE_CATEGORIAS_PROPIAS_GRATIS} "
                        f"categorías propias del plan gratis."
                    )
                else:
                    try:
                        insert_categoria(user.id, n, tipo_sel)
                        st.success(f"Categoría '{n}' creada")
                        st.cache_data.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")
    else:
        st.markdown("##### Agregar categoría")
        st.info(
            f"Llegaste al límite de {LIMITE_CATEGORIAS_PROPIAS_GRATIS} categorías "
            f"propias del plan gratis. Con **Premium** son ilimitadas. {MSG_PREMIUM}"
        )

    st.divider()

    cats_todas = get_categorias(user.id, tipo_sel, solo_activas=False)
    if not cats_todas:
        st.info(f"No hay categorías cargadas de tipo {tipo_sel}.")
        return

    cats_activas = [c for c in cats_todas if c["activa"]]
    cats_inactivas = [c for c in cats_todas if not c["activa"]]
    conteos = conteo_movs_por_categoria(user.id, tipo_sel)

    st.markdown(f"##### Activas ({len(cats_activas)})")
    if not cats_activas:
        st.caption("No hay categorías activas.")
    for c in cats_activas:
        _row_categoria(user, c, activa=True, conteos=conteos, premium=premium)

    if cats_inactivas:
        with st.expander(f"Inactivas ({len(cats_inactivas)})"):
            for c in cats_inactivas:
                _row_categoria(user, c, activa=False, conteos=conteos, premium=premium)


def _row_categoria(user, c, activa, conteos, premium):
    cat_id = c["id"]
    nombre_actual = c["nombre"]
    tipo_actual = c["tipo"]

    edit_key = f"edit_cat_{cat_id}"
    if st.session_state.get(edit_key):
        with st.form(f"form_edit_cat_{cat_id}"):
            col1, col2, col3 = st.columns([4, 1, 1])
            with col1:
                nuevo = st.text_input(
                    "Nuevo nombre", value=nombre_actual,
                    key=f"new_name_{cat_id}", label_visibility="collapsed",
                )
            with col2:
                guardar = st.form_submit_button("💾", use_container_width=True)
            with col3:
                cancelar = st.form_submit_button("✖", use_container_width=True)

            if guardar:
                n = (nuevo or "").strip()
                if not n:
                    st.error("El nombre no puede quedar vacío")
                elif n == nombre_actual:
                    st.session_state.pop(edit_key, None)
                    st.rerun()
                elif categoria_existe(user.id, n, tipo_actual, excluir_id=cat_id):
                    st.error(f"Ya existe '{n}' para tipo {tipo_actual}")
                else:
                    afectados = contar_movimientos_categoria(user.id, nombre_actual, tipo_actual)
                    try:
                        update_categoria(user.id, cat_id, nombre=n)
                        if afectados > 0:
                            renombrar_categoria_en_movimientos(user.id, nombre_actual, n)
                            st.success(
                                f"✅ Renombrada a '{n}'. {afectados} movimiento(s) actualizados."
                            )
                        else:
                            st.success(f"Renombrada a '{n}'")
                        st.session_state.pop(edit_key, None)
                        st.cache_data.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")
            if cancelar:
                st.session_state.pop(edit_key, None)
                st.rerun()
    else:
        col1, col2, col3 = st.columns([4, 1, 1])
        with col1:
            afectados = conteos.get(nombre_actual, 0)
            badge = (f"  <span style='color:{TEXT_MUTED};font-size:0.8em;'>"
                     f"({afectados} mov.)</span>") if afectados > 0 else ""
            st.markdown(
                f"<div style='padding-top:8px;'><strong style='color:{NAVY};'>{nombre_actual}</strong>{badge}</div>",
                unsafe_allow_html=True,
            )
        with col2:
            if st.button("✏️", key=f"btn_edit_{cat_id}", help="Renombrar",
                         use_container_width=True):
                if premium:
                    st.session_state[edit_key] = True
                    st.rerun()
                else:
                    st.info(MSG_PREMIUM)
        with col3:
            if activa:
                if st.button("🚫", key=f"btn_desact_{cat_id}", help="Desactivar",
                             use_container_width=True):
                    if not premium:
                        st.info(MSG_PREMIUM)
                    else:
                        try:
                            update_categoria(user.id, cat_id, activa=False)
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")
            else:
                if st.button("✅", key=f"btn_react_{cat_id}", help="Reactivar",
                             use_container_width=True):
                    if not premium:
                        st.info(MSG_PREMIUM)
                    else:
                        try:
                            update_categoria(user.id, cat_id, activa=True)
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")

# ============================================================================
# PANTALLA: INICIO (DEFAULT AL LOGUEARSE)
# ============================================================================
def _nombre_de_usuario(user):
    """Nombre para mostrar: usa el que configuró el usuario
    (user_metadata.display_name); si no hay, lo deriva del email."""
    meta = getattr(user, "user_metadata", None) or {}
    nombre = (meta.get("display_name") or meta.get("nombre") or "").strip()
    if nombre:
        return nombre
    email = getattr(user, "email", "") or ""
    local = email.split("@")[0] if "@" in email else email
    # Si tiene puntos o guiones, capitalizamos cada parte
    parts = [p.capitalize() for p in local.replace(".", " ").replace("-", " ").split()]
    return " ".join(parts) if parts else "👋"

def page_inicio(user):
    nombre = _nombre_de_usuario(user)
    page_header(f"Hola, {nombre}", "Tus finanzas desde un solo lugar.")

    estado = calcular_estado_flujo(user.id)
    if estado is None:
        st.info(
            "**Es tu primera vez por acá.** Para empezar:\n\n"
            "1. Andá a **Cargar movimiento** y registrá un ingreso, gasto o ahorro.\n"
            "2. Si ya tenías dinero antes de empezar, abrí **Reportes → Flujo de Fondos** "
            "y conciliá tu saldo: le decís a la app cuánto tenés realmente y listo.\n"
            "3. Volvé acá para ver tu resumen mes a mes."
        )
        return

    _dashboard_body(user, estado)

# ============================================================================
# CUERPO DEL DASHBOARD (se muestra dentro de Inicio)
# ============================================================================
def _dashboard_body(user, estado):
    pivot_completo = estado["pivot_completo"]
    pivot_dev = estado["pivot_dev"]
    flujo_neto = estado["flujo_neto"]
    meses_todos = estado["meses_todos"]
    saldos_fin = estado["saldos_fin"]

    # --- Selector de mes (default = último con datos) ---
    meses_lbl = [m.strftime("%m/%Y") for m in meses_todos]
    mes_idx = st.selectbox(
        "Mes a analizar",
        options=list(range(len(meses_lbl))),
        format_func=lambda i: meses_lbl[i],
        index=_default_mes_idx(meses_todos),
        key="dash_mes_idx",
    )
    mes_sel = meses_todos[mes_idx]
    mes_lbl_sel = meses_lbl[mes_idx]

    # --- 4 KPIs del mes seleccionado (criterio devengado, coincide con Movimientos) ---
    def safe_get(pivot, tipo, mes):
        if tipo in pivot.index and mes in pivot.columns:
            return float(pivot.loc[tipo, mes])
        return 0.0

    ingresos_mes = safe_get(pivot_dev, "Ingreso", mes_sel)
    gastos_mes = (safe_get(pivot_dev, "Gasto Fijo", mes_sel)
                  + safe_get(pivot_dev, "Gasto Variable", mes_sel))
    ahorro_mes = safe_get(pivot_dev, "Ahorro", mes_sel)
    resultado_mes = ingresos_mes - gastos_mes  # el Ahorro no es un gasto: se muestra aparte

    st.markdown(f"##### Mes: {mes_lbl_sel}")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(metric_card("Ingresos", fmt_money(ingresos_mes)),
                    unsafe_allow_html=True)
    with col2:
        st.markdown(metric_card("Gastos", fmt_money(gastos_mes), "orange"),
                    unsafe_allow_html=True)
    with col3:
        color_res = "green" if resultado_mes >= 0 else "red"
        st.markdown(metric_card("Resultado", fmt_money(resultado_mes), color_res,
                                sublabel="antes de ahorro"),
                    unsafe_allow_html=True)
    with col4:
        st.markdown(metric_card("Ahorro", fmt_money(ahorro_mes), "cyan"),
                    unsafe_allow_html=True)

    st.divider()

    # ------------------------------------------------------------------------
    # DISPONIBLE REAL (independiente del mes seleccionado arriba)
    # Saldo a fin del mes en curso vs cuotas de gastos ya comprometidas a futuro.
    # ------------------------------------------------------------------------
    hoy_d = date.today()
    mes_curso_ts = pd.Timestamp(hoy_d.year, hoy_d.month, 1)
    df_caja_all = estado["df_caja"]
    comprometido = float(df_caja_all[
        (df_caja_all["mes_pago"] > mes_curso_ts) &
        (df_caja_all["tipo"].isin(["Gasto Fijo", "Gasto Variable"]))
    ]["monto_cuota"].sum())

    saldo_hoy = None
    for m_ts in meses_todos:
        if pd.Timestamp(m_ts) <= mes_curso_ts:
            saldo_hoy = float(saldos_fin[m_ts])
        else:
            break

    if saldo_hoy is not None:
        st.markdown("##### Disponible real")
        cd1, cd2, cd3 = st.columns(3)
        with cd1:
            st.markdown(metric_card(f"Saldo a fin de {mes_curso_ts.strftime('%m/%Y')}",
                                    fmt_money(saldo_hoy)), unsafe_allow_html=True)
        with cd2:
            st.markdown(metric_card("Comprometido en cuotas futuras",
                                    fmt_money(comprometido), "orange"), unsafe_allow_html=True)
        with cd3:
            neto = saldo_hoy - comprometido
            st.markdown(metric_card("Disponible neto", fmt_money(neto),
                                    "" if neto >= 0 else "red"), unsafe_allow_html=True)
        st.caption(
            "El monto comprometido en cuotas futuras representa las cuotas que vencen "
            "en meses futuros respecto al actual. Una vez que pasa el mes, ese monto se "
            "actualiza asumiendo que las cuotas del período fueron abonadas."
        )
        st.divider()

    # ------------------------------------------------------------------------
    # RITMO DE GASTO VARIABLE (mes en curso)
    # ------------------------------------------------------------------------
    dias_transc = hoy_d.day
    dias_mes = pd.Timestamp(hoy_d).days_in_month
    gv_curso = safe_get(pivot_dev, "Gasto Variable", mes_curso_ts)
    if gv_curso > 0:
        st.markdown("##### Ritmo de gasto variable (mes en curso)")
        rg1, rg2, rg3 = st.columns(3)
        with rg1:
            st.markdown(metric_card("Gastado al día de hoy", fmt_money(gv_curso)),
                        unsafe_allow_html=True)
        with rg2:
            st.markdown(metric_card("Promedio diario", fmt_money(gv_curso / dias_transc),
                                    "orange"), unsafe_allow_html=True)
        with rg3:
            st.markdown(metric_card(f"Proyectado a fin de {mes_curso_ts.strftime('%m/%Y')}",
                                    fmt_money(gv_curso / dias_transc * dias_mes), "orange"),
                        unsafe_allow_html=True)
        st.caption(
            f"Promedio = gasto variable acumulado dividido los {dias_transc} días "
            f"transcurridos. El proyectado asume que seguís a este ritmo."
        )
        st.divider()

    # ------------------------------------------------------------------------
    # GRÁFICO 1: Composición por categoría (gastos + ahorro) con filtros propios
    # ------------------------------------------------------------------------
    st.markdown("##### Composición por categoría")

    # Selectores propios para este gráfico (independientes del selector principal)
    col_t1, col_t2 = st.columns([3, 2])
    with col_t1:
        # Default = mes principal del Dashboard; opción extra "Todos los meses"
        meses_lbl_torta = ["Todos los meses"] + meses_lbl
        idx_default = meses_lbl_torta.index(mes_lbl_sel) if mes_lbl_sel in meses_lbl_torta else 1
        mes_torta_sel = st.selectbox(
            "Mes",
            options=meses_lbl_torta,
            index=idx_default,
            key="dash_torta_mes",
        )
    with col_t2:
        tipo_torta_sel = st.selectbox(
            "Tipo",
            options=["Todos", "Gasto Fijo", "Gasto Variable", "Ahorro"],
            index=0,
            key="dash_torta_tipo",
        )

    df_dev = estado["df_dev"]

    # Filtrar por tipo
    if tipo_torta_sel == "Todos":
        df_torta_base = df_dev[df_dev["tipo"].isin(["Gasto Fijo", "Gasto Variable", "Ahorro"])].copy()
    else:
        df_torta_base = df_dev[df_dev["tipo"] == tipo_torta_sel].copy()

    # Filtrar por mes (criterio devengado: mes de la fecha de devengo)
    if mes_torta_sel != "Todos los meses":
        df_torta_base = df_torta_base[
            df_torta_base["fecha_devengo"].dt.strftime("%m/%Y") == mes_torta_sel
        ]

    if df_torta_base.empty:
        st.caption("No hay datos en la selección.")
    else:
        df_torta = (df_torta_base.groupby("categoria", as_index=False)["monto_total"].sum()
                    .sort_values("monto_total", ascending=False))
        total_torta = df_torta["monto_total"].sum()
        df_torta["monto_fmt"] = df_torta["monto_total"].apply(fmt_money)
        paleta = [NAVY, CYAN, ORANGE, GREEN, GREY,
                  NAVY_HOVER, "#3aa9c9", "#f08259", "#3fa85a", "#9aa0a6"]
        fig_torta = px.pie(
            df_torta, names="categoria", values="monto_total",
            hole=0.55,
            color_discrete_sequence=paleta,
            custom_data=["monto_fmt"],
        )
        fig_torta.update_traces(
            textposition="outside",
            textinfo="label+percent",
            automargin=True,
            marker=dict(line=dict(color="white", width=2)),
            hovertemplate="<b>%{label}</b><br>%{customdata[0]}<br>%{percent}<extra></extra>",
        )
        fig_torta = aplicar_tema_plotly(fig_torta, height=560)
        fig_torta.update_layout(margin=dict(t=55, b=55, l=70, r=70))
        fig_torta.update_layout(showlegend=False)
        st.plotly_chart(fig_torta, use_container_width=True, config={"displayModeBar": False})
        st.caption(
            f"Total {tipo_torta_sel.lower() if tipo_torta_sel != 'Todos' else 'gastos + ahorro'} "
            f"en {mes_torta_sel.lower() if mes_torta_sel != 'Todos los meses' else 'todos los meses'}: "
            f"**{fmt_money(total_torta)}**"
        )


# ============================================================================
# APP PRINCIPAL (LOGUEADO)
# ============================================================================
def app(user):
    # Header minimalista: lockup de marca (símbolo embebido + nombre) y cuenta a la derecha.
    col_brand, col_acc = st.columns([3, 2])
    with col_brand:
        st.markdown(
            f"<div style='display:flex; align-items:center; gap:11px; padding-top:2px;'>"
            f"<img src='{_LOGO_DATA_URI}' width='44' height='44' style='display:block; flex:0 0 auto;'/>"
            f"<div style='line-height:1.1;'>"
            f"<div style='font-family:Poppins,sans-serif; font-weight:700; font-size:1.15rem; "
            f"color:{NAVY}; letter-spacing:-0.01em;'>Finanzas <span style='color:{CYAN};'>WL</span></div>"
            f"<div style='font-size:0.62rem; color:{TEXT_MUTED}; letter-spacing:0.1em; "
            f"text-transform:uppercase; margin-top:1px;'>WL HNOS &amp; ASOC</div>"
            f"</div></div>",
            unsafe_allow_html=True,
        )
    with col_acc:
        if st.button("Cerrar sesión", use_container_width=True, key="logout_top"):
            do_signout()
        st.markdown(
            f"<div style='text-align:right; color:{TEXT_MUTED}; font-size:0.72rem; "
            f"margin-top:4px; word-break:break-all;'>{user.email}</div>",
            unsafe_allow_html=True,
        )

    # Navegación principal, siempre visible en cualquier pantalla.
    page = st.radio("Menú", [
        "Inicio",
        "Cargar",
        "Movimientos",
        "Reportes",
        "Configuración",
    ], horizontal=True, label_visibility="collapsed", key="nav_principal")
    st.divider()

    if "Inicio" in page:
        page_inicio(user)
    elif "Cargar" in page:
        page_cargar(user)
    elif "Movimientos" in page:
        page_ver(user)
    elif "Reportes" in page:
        page_reportes(user)
    elif "Configuración" in page:
        page_configuracion(user)

    # Footer de marca con link a la landing
    st.markdown(
        f"<div style='text-align:center; margin-top:42px; padding-top:14px; "
        f"border-top:1px solid {BORDER}; color:{TEXT_MUTED}; font-size:0.8rem;'>"
        f"<a href='https://wlhnos.vercel.app/' target='_blank' "
        f"style='color:{CYAN}; text-decoration:none; font-weight:600;'>WL HNOS &amp; ASOC</a>"
        f" · Catamarca</div>",
        unsafe_allow_html=True,
    )

# ============================================================================
# ROUTER
# ============================================================================
# Vuelta del login con Google: PKCE, el código llega en ?code= (legible por el servidor).
_qp = st.query_params
if _qp.get("error_description"):
    st.session_state["oauth_error"] = _qp.get("error_description")
    st.query_params.clear()
    st.rerun()

if _qp.get("code"):
    _verifier = None
    try:
        _verifier = cookie_manager.get("wl_pkce")
    except Exception:
        pass
    if not _verifier:
        _verifier = st.session_state.get("pkce_verifier")

    if not _verifier and st.session_state.get("pkce_intentos", 0) < 2:
        # La cookie todavía no llegó del navegador (primer render): esperar
        # un ciclo; el componente de cookies dispara un rerun solo.
        st.session_state["pkce_intentos"] = st.session_state.get("pkce_intentos", 0) + 1
        st.stop()

    if _verifier:
        try:
            resp = httpx.post(
                f"{st.secrets['SUPABASE_URL'].strip().rstrip('/')}/auth/v1/token?grant_type=pkce",
                json={"auth_code": _qp.get("code"), "code_verifier": _verifier},
                headers={"apikey": st.secrets["SUPABASE_KEY"],
                         "Content-Type": "application/json"},
                timeout=15,
            )
            data = resp.json()
            if resp.status_code == 200 and data.get("access_token"):
                sb.auth.set_session(data["access_token"], data["refresh_token"])
                u = sb.auth.get_user()
                if u and u.user:
                    st.session_state.sb_tokens = {
                        "access_token": data["access_token"],
                        "refresh_token": data["refresh_token"],
                    }
                    st.session_state.user = u.user
                    _guardar_cookie_sesion(data["refresh_token"], "ck_oauth")
                    try:
                        seed_categorias_si_vacio(u.user.id)
                    except Exception:
                        pass
                else:
                    st.session_state["oauth_error"] = "no se pudo leer el usuario."
            else:
                st.session_state["oauth_error"] = (
                    data.get("error_description") or data.get("msg") or str(data)
                )
        except Exception as e:
            st.session_state["oauth_error"] = str(e)
    else:
        st.session_state["oauth_error"] = (
            "no llegó el código de verificación del navegador. Probá de nuevo."
        )
    st.session_state["pkce_intentos"] = 0
    st.session_state["pkce_verifier"] = None
    try:
        cookie_manager.delete("wl_pkce", key="ck_pkce_del")
    except Exception:
        pass
    st.query_params.clear()
    st.rerun()

# Si la URL trae el token del mail de recuperación, capturarlo y limpiar la URL.
if _qp.get("type") == "recovery" and _qp.get("token_hash"):
    st.session_state.recovery_token_hash = _qp.get("token_hash")
    st.query_params.clear()

if st.session_state.get("recovery_token_hash"):
    page_nueva_password()
elif st.session_state.user is None:
    page_login()
else:
    app(st.session_state.user)
